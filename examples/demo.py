"""XLStruct demo — schema-driven Excel extraction with LLM.

Covers: basic extraction, ExtractionConfig, report, provenance, DataFrame export.

Usage:
    uv run python examples/demo.py

Requires:
    - API key in scripts/.env (OPENAI_API_KEY or ANTHROPIC_API_KEY)
    - pandas: uv add pandas
"""

import asyncio
import tempfile
from pathlib import Path as PathLibPath

import openpyxl
from dotenv import load_dotenv
from pydantic import BaseModel, Field

load_dotenv("scripts/.env")

from xlstruct.config import ExtractionConfig, ExtractionMode
from xlstruct.extractor import Extractor


# * Schemas

class Product(BaseModel):
    name: str
    category: str
    price: float
    stock: int


class Employee(BaseModel):
    full_name: str = Field(description="First and last name combined")
    department: str
    annual_salary: float = Field(description="Annual salary in USD")
    start_date: str = Field(description="ISO format date (YYYY-MM-DD)")
    is_active: bool = Field(description="Whether employee is currently active")


# * Sample data generators

def create_product_xlsx() -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    ws["A1"] = "Product Name"
    ws["B1"] = "Category"
    ws["C1"] = "Price"
    ws["D1"] = "Stock"

    data = [
        ("MacBook Pro 14", "Laptop", 1999.00, 50),
        ("Dell XPS 15", "Laptop", 1499.00, 30),
        ("iPhone 16 Pro", "Phone", 1199.00, 200),
        ("Galaxy S25", "Phone", 999.00, 150),
        ("iPad Air", "Tablet", 599.00, 80),
        ("AirPods Pro", "Accessory", 249.00, 500),
    ]
    for i, (name, cat, price, stock) in enumerate(data, start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = cat
        ws[f"C{i}"] = price
        ws[f"D{i}"] = stock

    path = PathLibPath(tempfile.mkdtemp()) / "products.xlsx"
    wb.save(str(path))
    return str(path)


def create_employee_xlsx() -> str:
    """Messy spreadsheet: merged title, inconsistent dates, monthly salary."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    ws.merge_cells("A1:F1")
    ws["A1"] = "ACME Corp — Employee Directory"

    ws["A3"] = "First"
    ws["B3"] = "Last"
    ws["C3"] = "Dept."
    ws["D3"] = "Salary (Monthly)"
    ws["E3"] = "Joined"
    ws["F3"] = "Status"

    data = [
        ("John", "Doe", "Engineering", 8500, "2022-03-15", "Active"),
        ("Jane", "Smith", "Marketing", 7200, "Jan 2023", "Active"),
        ("Bob", "Wilson", "Engineering", 9000, "2021/11/01", "On Leave"),
        ("Alice", "Brown", "Sales", 6800, "2023-06-01", "Terminated"),
        ("Charlie", "Davis", "Engineering", 8800, "15-Mar-2020", "Active"),
    ]
    for i, row in enumerate(data, start=4):
        for j, val in enumerate(row):
            ws.cell(row=i, column=j + 1, value=val)

    path = PathLibPath(tempfile.mkdtemp()) / "employees.xlsx"
    wb.save(str(path))
    return str(path)


# * Demo

async def main():
    extractor = Extractor()  # ^ Default: anthropic/claude-sonnet-4-6

    # * 1. Basic extraction
    print("=" * 60)
    print("1. Basic extraction")
    print("=" * 60)
    product_path = create_product_xlsx()
    result = await extractor.extract(product_path, Product)
    print(result.report)
    print()
    for item in result:
        print(f"   {item.name} ({item.category}): ${item.price:.2f}")

    # * 2. DataFrame export
    print(f"\n{'=' * 60}")
    print("2. to_dataframe()")
    print("=" * 60)
    df = result.to_dataframe()
    print(df.to_string(index=False))

    # * 3. Provenance tracking
    print(f"\n{'=' * 60}")
    print("3. Row provenance")
    print("=" * 60)
    config = ExtractionConfig(
        output_schema=Product,
        header_rows=[1],
        track_provenance=True,
    )
    result2 = await extractor.extract(product_path, extraction_config=config)
    print(result2.report)
    print()
    for item, rows in zip(result2, result2.report.source_rows):
        print(f"   {item.name} <- Row {rows}")

    # * 4. Custom instructions (messy spreadsheet)
    print(f"\n{'=' * 60}")
    print("4. Custom instructions (messy spreadsheet)")
    print("=" * 60)
    employee_path = create_employee_xlsx()
    result3 = await extractor.extract(
        employee_path,
        Employee,
        instructions=(
            "- Combine First and Last name columns into full_name\n"
            "- Salary column is MONTHLY — multiply by 12 for annual_salary\n"
            "- Normalize all dates to YYYY-MM-DD\n"
            "- is_active: only 'Active' = True, everything else = False\n"
            "- Skip rows where status is 'Terminated'"
        ),
    )
    print(result3.report)
    print()
    for emp in result3:
        print(
            f"   {emp.full_name} | {emp.department} | "
            f"${emp.annual_salary:,.0f}/yr | {emp.start_date} | "
            f"active={emp.is_active}"
        )

    # * 5. Mode selection
    print(f"\n{'=' * 60}")
    print("5. Explicit mode selection")
    print("=" * 60)
    for mode in [ExtractionMode.DIRECT, ExtractionMode.AUTO]:
        config = ExtractionConfig(
            output_schema=Product,
            mode=mode,
            header_rows=[1],
        )
        r = await extractor.extract(product_path, extraction_config=config)
        print(f"   mode={mode.value:8s} -> resolved={r.report.mode.value}")


if __name__ == "__main__":
    asyncio.run(main())
