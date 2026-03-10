"""Basic extraction example: Parse an Excel file into Pydantic models.

Usage:
    uv run python examples/basic_extraction.py

Requires:
    - An OpenAI API key set as OPENAI_API_KEY env var
    - A sample Excel file (created automatically by this script)
"""

import asyncio
import tempfile
from pathlib import Path as PathLibPath

import openpyxl
from pydantic import BaseModel

from xlstruct.extractor import Extractor


# * Define your target schema
class InvoiceItem(BaseModel):
    item_name: str
    quantity: int
    unit_price: float
    total: float


def create_sample_xlsx() -> str:
    """Create a sample invoice Excel file for demonstration."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ^ Header
    ws["A1"] = "Item"
    ws["B1"] = "Description"
    ws["C1"] = "Qty"
    ws["D1"] = "Unit Price"
    ws["E1"] = "Total"

    # ^ Data
    items = [
        ("WDG-001", "Widget Alpha", 10, 25.00, 250.00),
        ("WDG-002", "Widget Beta", 5, 42.50, 212.50),
        ("WDG-003", "Widget Gamma", 20, 15.75, 315.00),
    ]
    for i, (code, desc, qty, price, total) in enumerate(items, start=2):
        ws[f"A{i}"] = code
        ws[f"B{i}"] = desc
        ws[f"C{i}"] = qty
        ws[f"D{i}"] = price
        ws[f"E{i}"] = total

    path = PathLibPath(tempfile.mkdtemp()) / "invoice.xlsx"
    wb.save(str(path))
    return str(path)


async def main():
    # * Create sample file
    xlsx_path = create_sample_xlsx()
    print(f"Sample file: {xlsx_path}")

    # * Create Extractor instance
    extractor = Extractor(provider="openai/gpt-4o")

    # * Extract structured data
    items = await extractor.extract(
        xlsx_path,
        InvoiceItem,
        instructions="Extract invoice line items. Use the Total column value.",
    )

    # * Print results
    print(f"\nExtracted {len(items)} items:")
    for item in items:
        print(f"  {item.item_name}: {item.quantity} x ${item.unit_price:.2f} = ${item.total:.2f}")


if __name__ == "__main__":
    asyncio.run(main())
