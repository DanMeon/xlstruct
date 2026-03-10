"""Custom instructions example: Guide LLM behavior with natural language.

Usage:
    uv run python examples/custom_instructions.py

Shows how to use the `instructions` parameter to steer LLM extraction
for complex or ambiguous spreadsheets.
"""

import asyncio
import tempfile
from pathlib import Path as PathLibPath

import openpyxl
from pydantic import BaseModel, Field

from xlstruct.extractor import Extractor


# * Schema with field-level descriptions
class EmployeeRecord(BaseModel):
    full_name: str = Field(description="First and last name combined")
    department: str
    annual_salary: float = Field(description="Annual salary in USD")
    start_date: str = Field(description="ISO format date (YYYY-MM-DD)")
    is_active: bool = Field(description="Whether employee is currently active")


def create_messy_xlsx() -> str:
    """Create a deliberately messy spreadsheet to demonstrate instruction power."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # ^ Merged title row
    ws.merge_cells("A1:F1")
    ws["A1"] = "ACME Corp — Employee Directory (Confidential)"

    # ^ Empty row
    # Row 2 is blank

    # ^ Column headers (row 3) — inconsistent naming
    ws["A3"] = "First"
    ws["B3"] = "Last"
    ws["C3"] = "Dept."
    ws["D3"] = "Salary (Monthly)"  # ^ Note: monthly, not annual
    ws["E3"] = "Joined"
    ws["F3"] = "Status"

    # ^ Data (starting row 4) — mixed formats
    data = [
        ("John", "Doe", "Engineering", 8500, "2022-03-15", "Active"),
        ("Jane", "Smith", "Marketing", 7200, "Jan 2023", "Active"),
        ("Bob", "Wilson", "Engineering", 9000, "2021/11/01", "On Leave"),
        ("Alice", "Brown", "Sales", 6800, "2023-06-01", "Terminated"),
        ("Charlie", "Davis", "Engineering", 8800, "15-Mar-2020", "Active"),
    ]
    for i, (first, last, dept, salary, joined, status) in enumerate(data, start=4):
        ws[f"A{i}"] = first
        ws[f"B{i}"] = last
        ws[f"C{i}"] = dept
        ws[f"D{i}"] = salary
        ws[f"E{i}"] = joined
        ws[f"F{i}"] = status

    path = PathLibPath(tempfile.mkdtemp()) / "employees.xlsx"
    wb.save(str(path))
    return str(path)


async def main():
    xlsx_path = create_messy_xlsx()
    print(f"Sample file: {xlsx_path}")

    extractor = Extractor(provider="openai/gpt-4o")

    # * Custom instructions handle the messy data
    instructions = """
    - Combine First and Last name columns into full_name
    - The Salary column shows MONTHLY salary — multiply by 12 for annual_salary
    - Normalize all dates to ISO format (YYYY-MM-DD)
    - For is_active: only "Active" status = True, everything else = False
    - Skip any rows where status is "Terminated"
    """

    records = await extractor.extract(
        xlsx_path,
        EmployeeRecord,
        instructions=instructions,
    )

    print(f"\nExtracted {len(records)} active/on-leave employees:")
    for emp in records:
        print(
            f"  {emp.full_name} | {emp.department} | "
            f"${emp.annual_salary:,.0f}/yr | {emp.start_date} | "
            f"Active: {emp.is_active}"
        )


if __name__ == "__main__":
    asyncio.run(main())
