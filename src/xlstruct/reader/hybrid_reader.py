"""HybridReader: calamine + openpyxl hybrid Excel parser.

Pass 1 (calamine, Rust): values, merged cells, data types, dimensions
Pass 2 (openpyxl, read-only): formula strings only — skipped for .xls

Supports .xlsx, .xlsm, .xltx, .xltm, .xls, .ods.
Requires python-calamine >= 0.6.2.
"""

import io
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import Any

import openpyxl
import python_calamine as pc
from openpyxl.utils import get_column_letter

from xlstruct.exceptions import ErrorCode, ReaderError
from xlstruct.schemas.core import CellData, SheetData, WorkbookData


@dataclass
class _CalamineSheetData:
    """Intermediate storage for hybrid 2-pass data accumulation."""

    name: str
    dimensions: str = ""
    row_count: int = 0
    col_count: int = 0
    merged_ranges: list[str] = field(default_factory=list)
    # ^ {(1-indexed row, col): range_str}
    merged_cell_map: dict[tuple[int, int], str] = field(default_factory=dict)
    # ^ {(1-indexed row, col): (1-indexed origin row, col)}
    merge_origins: dict[tuple[int, int], tuple[int, int]] = field(default_factory=dict)
    # ^ {(1-indexed row, col): calamine value}
    values: dict[tuple[int, int], Any] = field(default_factory=dict)
    # ^ {(1-indexed row, col): "s"|"n"|"d"|"b"}
    data_types: dict[tuple[int, int], str] = field(default_factory=dict)
    # ^ {(1-indexed row, col): formula_string} — populated by openpyxl pass
    formulas: dict[tuple[int, int], str] = field(default_factory=dict)
    # ^ {(1-indexed row, col): number_format_string} — populated by openpyxl pass
    number_formats: dict[tuple[int, int], str] = field(default_factory=dict)

    # ^ Formats where openpyxl can extract formula strings (Pass 2)


_OPENPYXL_FORMATS = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})


class HybridReader:
    """Excel reader using calamine (Rust) + openpyxl hybrid pattern.

    Pass 1 (calamine): values, merged cells, data types, dimensions — all formats
    Pass 2 (openpyxl, read_only=True, data_only=False): formula strings — .xlsx/.xlsm only

    For .xls files, Pass 2 is skipped (openpyxl doesn't support .xls).
    Formula strings are unavailable for .xls but computed values are preserved.
    """

    def read(
        self,
        file_bytes: bytes,
        sheet_name: str | None = None,
        *,
        source_ext: str = ".xlsx",
    ) -> WorkbookData:
        """Read an Excel file from bytes into WorkbookData.

        Args:
            file_bytes: Raw bytes of the Excel file.
            sheet_name: If provided, read only this sheet. None = all sheets.
            source_ext: File extension (e.g. ".xlsx", ".xls") to determine
                whether Pass 2 (formula extraction) is possible.
        """
        buf = io.BytesIO(file_bytes)

        try:
            cwb = pc.CalamineWorkbook.from_filelike(buf)
            all_names = cwb.sheet_names
        except Exception as e:
            raise ReaderError(
                f"Failed to open workbook: {e}", code=ErrorCode.READER_PARSE_FAILED
            ) from e

        if sheet_name is not None:
            if sheet_name not in all_names:
                raise ReaderError(
                    f"Sheet '{sheet_name}' not found. Available: {all_names}",
                    code=ErrorCode.READER_PARSE_FAILED,
                )
            target_sheets = [sheet_name]
        else:
            target_sheets = list(all_names)

        try:
            # * Pass 1: calamine — values, merged cells, types, dimensions
            calamine_data = self._calamine_pass(cwb, target_sheets)

            # * Pass 2: openpyxl — formula strings (only for supported formats)
            if source_ext.lower() in _OPENPYXL_FORMATS:
                buf.seek(0)
                self._formula_pass(buf, target_sheets, calamine_data)

        except ReaderError:
            raise
        except Exception as e:
            raise ReaderError(
                f"Failed to read workbook: {e}", code=ErrorCode.READER_PARSE_FAILED
            ) from e

        # * Build final WorkbookData
        sheets: list[SheetData] = []
        for sn in target_sheets:
            sheet = self._build_sheet_data(calamine_data[sn])
            # * Reject formula cells without cached values
            self._check_uncached_formulas(sheet)
            sheets.append(sheet)
        return WorkbookData(sheets=sheets)

    def _calamine_pass(
        self,
        cwb: pc.CalamineWorkbook,
        target_sheets: list[str],
    ) -> dict[str, _CalamineSheetData]:
        """Pass 1: Extract values, merged cells, types, dimensions via calamine."""
        result: dict[str, _CalamineSheetData] = {}

        for sn in target_sheets:
            sheet = cwb.get_sheet_by_name(sn)
            cal = _CalamineSheetData(
                name=sn,
                row_count=sheet.height,
                col_count=sheet.width,
            )

            # * Dimensions
            if sheet.height > 0 and sheet.width > 0:
                cal.dimensions = f"A1:{get_column_letter(sheet.width)}{sheet.height}"

            # * Merged cells (calamine: 0-indexed → 1-indexed)
            merged = sheet.merged_cell_ranges
            if merged is not None:
                for (r_start, c_start), (r_end, c_end) in merged:
                    range_str = self._calamine_merge_to_range_str(
                        (r_start, c_start), (r_end, c_end)
                    )
                    cal.merged_ranges.append(range_str)

                    origin = (r_start + 1, c_start + 1)
                    for r in range(r_start, r_end + 1):
                        for c in range(c_start, c_end + 1):
                            key = (r + 1, c + 1)
                            cal.merged_cell_map[key] = range_str
                            if key != origin:
                                cal.merge_origins[key] = origin

            # * Values and data types
            rows = sheet.to_python()
            for r_idx, row in enumerate(rows):
                for c_idx, value in enumerate(row):
                    key = (r_idx + 1, c_idx + 1)

                    # ^ Skip empty cells (calamine returns '' for empty)
                    if value == "" or value is None:
                        continue

                    cal.data_types[key] = self._infer_data_type(value)
                    # ^ Convert date/datetime to ISO string for CellData compatibility
                    if isinstance(value, (datetime, date, time, timedelta)):
                        cal.values[key] = (
                            value.isoformat() if hasattr(value, "isoformat") else str(value)
                        )
                    else:
                        cal.values[key] = value

            result[sn] = cal

        return result

    def _formula_pass(
        self,
        buf: io.BytesIO,
        target_sheets: list[str],
        calamine_data: dict[str, _CalamineSheetData],
    ) -> None:
        """Pass 2: Extract formula strings and number formats from openpyxl (read_only mode)."""
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=False)
        try:
            for sn in target_sheets:
                ws = wb[sn]
                ws.reset_dimensions()
                cal = calamine_data[sn]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        val = cell.value
                        if isinstance(val, str) and val.startswith("="):
                            cal.formulas[(cell.row, cell.column)] = val
                        # ^ Store number format if it's not the uninformative default
                        nf = cell.number_format
                        if nf and nf != "General":
                            cal.number_formats[(cell.row, cell.column)] = nf
        finally:
            wb.close()

    def _build_sheet_data(self, cal: _CalamineSheetData) -> SheetData:
        """Merge calamine values and openpyxl formulas into SheetData."""
        # ^ Union of all cell positions from both passes
        all_keys = set(cal.values.keys()) | set(cal.formulas.keys())

        cells: list[CellData] = []
        for row, col in sorted(all_keys):
            value = cal.values.get((row, col))
            data_type = cal.data_types.get((row, col), "n")
            formula = cal.formulas.get((row, col))
            number_format = cal.number_formats.get((row, col))
            is_merged = (row, col) in cal.merged_cell_map
            merge_range = cal.merged_cell_map.get((row, col))
            merge_origin = cal.merge_origins.get((row, col))

            if formula:
                # ^ Formula cell: value = formula string, cached_value = computed value
                cells.append(
                    CellData(
                        row=row,
                        col=col,
                        value=formula,
                        formula=formula,
                        cached_value=value,
                        data_type="f",
                        number_format=number_format,
                        is_merged=is_merged,
                        merge_range=merge_range,
                        merge_origin=merge_origin,
                    )
                )
            else:
                # ^ Normal cell: value = calamine value, cached_value = same
                cells.append(
                    CellData(
                        row=row,
                        col=col,
                        value=value,
                        formula=None,
                        cached_value=value,
                        data_type=data_type,
                        number_format=number_format,
                        is_merged=is_merged,
                        merge_range=merge_range,
                        merge_origin=merge_origin,
                    )
                )

        return SheetData(
            name=cal.name,
            dimensions=cal.dimensions,
            cells=cells,
            merged_ranges=cal.merged_ranges,
            row_count=cal.row_count,
            col_count=cal.col_count,
        )

    @staticmethod
    def _check_uncached_formulas(sheet: SheetData) -> None:
        """Raise if any formula cells have no cached value."""
        uncached = [
            f"{get_column_letter(c.col)}{c.row}"
            for c in sheet.cells
            if c.formula and c.cached_value is None
        ]
        if uncached:
            sample = ", ".join(uncached[:5])
            suffix = f" (and {len(uncached) - 5} more)" if len(uncached) > 5 else ""
            raise ReaderError(
                f"Sheet '{sheet.name}': {len(uncached)} formula cell(s) have no cached value. "
                f"Open the file in Excel and re-save to populate cached values. "
                f"Cells: {sample}{suffix}",
                code=ErrorCode.READER_PARSE_FAILED,
            )

    @staticmethod
    def _infer_data_type(value: Any) -> str:
        """Map Python types from calamine to openpyxl-compatible data_type codes."""
        # ^ bool MUST be checked before int (bool is subclass of int in Python)
        if isinstance(value, bool):
            return "b"
        if isinstance(value, (int, float)):
            return "n"
        if isinstance(value, str):
            return "s"
        if isinstance(value, (datetime, date, time, timedelta)):
            return "d"
        return "n"

    @staticmethod
    def _calamine_merge_to_range_str(
        start: tuple[int, int],
        end: tuple[int, int],
    ) -> str:
        """Convert calamine 0-indexed merge coords to 'A1:C3' format."""
        r_start, c_start = start
        r_end, c_end = end
        return (
            f"{get_column_letter(c_start + 1)}{r_start + 1}"
            f":{get_column_letter(c_end + 1)}{r_end + 1}"
        )
