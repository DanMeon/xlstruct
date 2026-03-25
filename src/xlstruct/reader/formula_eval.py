"""Optional formula evaluation using the formulas library.

Builds a temporary .xlsx workbook from SheetData, evaluates formulas
via formulas.ExcelModel, and patches cached_value on formula cells.

Requires: pip install xlstruct[formulas]
"""

import logging
import tempfile
from pathlib import Path as PathLibPath
from typing import Any

from openpyxl.utils import get_column_letter

from xlstruct.schemas.core import CellData, SheetData

logger = logging.getLogger(__name__)


def _import_formulas() -> Any:
    """Lazy import of the formulas library with a helpful error message."""
    try:
        import formulas

        return formulas
    except ImportError:
        raise ImportError(
            "The 'formulas' package is required for formula evaluation. "
            "Install it with: pip install xlstruct[formulas]"
        ) from None


def evaluate_sheet_formulas(sheet: SheetData) -> SheetData:
    """Evaluate formula cells in a SheetData and update their cached_value.

    Builds a temporary .xlsx file from the sheet's cell data, runs the
    formulas ExcelModel to compute results, and patches each formula cell's
    cached_value with the computed result.

    Per-cell errors are caught individually so one failure does not prevent
    other cells from being evaluated.

    Args:
        sheet: The SheetData containing formula cells to evaluate.

    Returns:
        A new SheetData with updated cached_value on formula cells.
    """
    # * Collect formula cells
    formula_cells = [c for c in sheet.cells if c.formula]
    if not formula_cells:
        return sheet

    formulas_lib = _import_formulas()

    # * Build temporary .xlsx with all cell data
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet.name

    for cell in sheet.cells:
        coord = f"{get_column_letter(cell.col)}{cell.row}"
        if cell.formula:
            ws[coord] = cell.formula
        elif cell.value is not None:
            ws[coord] = cell.value

    # * Write to temp file (formulas requires a file path)
    tmp_path: PathLibPath | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = PathLibPath(tmp.name)
            wb.save(tmp.name)
        wb.close()

        # * Run formulas ExcelModel
        model = formulas_lib.ExcelModel().loads(str(tmp_path)).finish()
        solution = model.calculate()

        # * Extract computed values keyed by (row, col)
        computed: dict[tuple[int, int], Any] = {}
        for key, value in solution.items():
            # ^ key format: "'[book]Sheet'!CELL" — parse the cell reference
            try:
                cell_ref = str(key).rsplit("!", 1)[-1].strip("'\"")
                # ^ Use openpyxl to parse cell coordinate
                from openpyxl.utils.cell import (
                    column_index_from_string,
                    coordinate_from_string,
                )

                col_letter, row_num = coordinate_from_string(cell_ref)
                col_num = column_index_from_string(col_letter)

                # ^ formulas may return numpy arrays for single cells
                import numpy as np

                if isinstance(value, np.ndarray):
                    value = value.item() if value.size == 1 else value.flat[0]
                # ^ Convert numpy scalar types to Python native
                if isinstance(value, np.integer):
                    value = int(value)
                elif isinstance(value, np.floating):
                    value = float(value)
                elif isinstance(value, np.bool_):
                    value = bool(value)

                computed[(row_num, col_num)] = value
            except Exception:
                logger.warning("Failed to parse formula result key: %s", key)
                continue

    finally:
        if tmp_path is not None:
            tmp_path.unlink(missing_ok=True)

    # * Patch formula cells with computed values
    updated_cells: list[CellData] = []
    for cell in sheet.cells:
        if cell.formula and (cell.row, cell.col) in computed:
            new_value = computed[(cell.row, cell.col)]
            try:
                updated_cells.append(cell.model_copy(update={"cached_value": new_value}))
            except Exception:
                logger.warning(
                    "Failed to update cached_value for cell %s%d",
                    get_column_letter(cell.col),
                    cell.row,
                )
                updated_cells.append(cell)
        else:
            updated_cells.append(cell)

    return sheet.model_copy(update={"cells": updated_cells})
