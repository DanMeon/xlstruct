"""Shared encoding utilities for CompressedEncoder.

Centralizes: cell formatting, header detection, merged cell annotation,
formula summarization, and column type analysis.
"""

import math
import re
from collections import Counter
from datetime import date, datetime, time

from openpyxl.utils import get_column_letter, range_boundaries

from xlstruct.schemas.core import CellData, SheetData

# * Cell value formatting


def format_cell_value(cell: CellData) -> str:
    """Format a cell's display value as a clean string.

    Uses cached_value if available, else value.
    """
    val = cell.display_value
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, float):
        # ^ Remove trailing zeros: 250.00 → "250", 42.5 → "42.5"
        if not math.isfinite(val):
            return str(val)
        if val == int(val):  # pyright: ignore[reportUnnecessaryComparison]
            return str(int(val))
        return str(val)
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, time):
        return val.isoformat()
    return str(val)


# * Header detection


def detect_header_row(sheet: SheetData) -> int | None:
    """Estimate the header row number (1-indexed).

    Heuristics:
    1. First non-empty row after any merged header region
    2. Row with highest ratio of string-type cells
    3. Falls back to first non-empty row
    """
    if not sheet.cells:
        return None

    # ^ Find the max row of top merged regions (likely title/header area)
    # ^ Only multi-column merges count as title regions — a single-column
    # ^ merge (e.g., N1:N2) shouldn't disqualify an entire row as header
    top_merge_end = 0
    for mr_str in sheet.merged_ranges:
        match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", mr_str)
        if match:
            start_col, start_row = match.group(1), int(match.group(2))
            end_col, end_row = match.group(3), int(match.group(4))
            if start_row <= 5 and start_col != end_col:
                top_merge_end = max(top_merge_end, end_row)

    # ^ Group cells by row
    rows: dict[int, list[CellData]] = {}
    for cell in sheet.cells:
        if cell.merge_origin is not None:
            continue  # ^ Skip non-origin merged cells
        rows.setdefault(cell.row, []).append(cell)

    # ^ Find the row with highest string ratio after any top merges
    best_row: int | None = None
    best_score = -1.0

    for row_num in sorted(rows):
        if row_num <= top_merge_end:
            continue
        row_cells = rows[row_num]
        if not row_cells:
            continue

        non_empty = [c for c in row_cells if c.display_value is not None]
        if not non_empty:
            continue

        string_count = sum(
            1
            for c in non_empty
            if c.data_type in ("s", "str", "inlineStr") or isinstance(c.display_value, str)
        )
        score = string_count / len(non_empty)

        if score > best_score:
            best_score = score
            best_row = row_num

    # ^ Header should have at least 50% strings
    if best_row is not None and best_score >= 0.5:
        return best_row

    # ^ Fallback: first non-empty row
    for row_num in sorted(rows):
        non_empty = [c for c in rows[row_num] if c.display_value is not None]
        if non_empty:
            return row_num

    return None


def build_column_headers(
    sheet: SheetData,
    header_row: int,
) -> dict[int, str]:
    """Extract {col_index: header_name} mapping from the header row."""
    headers: dict[int, str] = {}
    for cell in sheet.cells:
        if cell.row == header_row and cell.display_value is not None:
            headers[cell.col] = str(cell.display_value).strip()
    return headers


def build_multi_row_headers(
    sheet: SheetData,
    header_rows: list[int],
) -> dict[int, str]:
    """Build combined column headers from multiple header rows.

    For merged cells, propagates the origin cell's value to all columns
    within the merge range. Combines values across rows with " / ".

    Example:
        Row 1: D1:E1="1 - 1" (merged)
        Row 2: D2="East", E2="West"
        → {4: "1 - 1 / East", 5: "1 - 1 / West"}
    """
    header_rows_set = set(header_rows)

    # * Build merge lookup: (row, col) → (origin_row, origin_col)
    merge_lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for range_str in sheet.merged_ranges:
        bounds = range_boundaries(range_str)
        # ^ range_boundaries returns Optional ints, but merge ranges always have all four
        assert bounds[0] is not None and bounds[1] is not None
        assert bounds[2] is not None and bounds[3] is not None
        min_col, min_row, max_col, max_row = bounds
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    merge_lookup[(r, c)] = (min_row, min_col)

    # * Build per-row header values
    per_row: list[dict[int, str]] = []
    for row_num in header_rows:
        row_headers: dict[int, str] = {}
        for col in range(1, sheet.col_count + 1):
            cell = sheet.get_cell(row_num, col)

            if cell is not None and cell.display_value is not None:
                row_headers[col] = str(cell.display_value).strip()
                continue

            # ^ Cell missing or empty — check merge lookup for origin
            origin = (
                cell.merge_origin
                if (cell is not None and cell.merge_origin is not None)
                else merge_lookup.get((row_num, col))
            )
            if origin is not None:
                origin_row, origin_col = origin
                if origin_row == row_num:
                    # ^ Horizontal merge (same row): propagate origin value
                    origin_cell = sheet.get_cell(origin_row, origin_col)
                    if origin_cell and origin_cell.display_value is not None:
                        row_headers[col] = str(origin_cell.display_value).strip()
                elif origin_row not in header_rows_set:
                    # ^ Vertical merge from outside headers: propagate
                    origin_cell = sheet.get_cell(origin_row, origin_col)
                    if origin_cell and origin_cell.display_value is not None:
                        row_headers[col] = str(origin_cell.display_value).strip()
                # ^ else: vertical merge within headers — skip (origin row handles it)

        per_row.append(row_headers)

    # * Combine across rows with " / "
    all_cols: set[int] = set()
    for rh in per_row:
        all_cols.update(rh.keys())

    combined: dict[int, str] = {}
    for col in sorted(all_cols):
        parts = [rh.get(col, "") for rh in per_row]
        non_empty = [p for p in parts if p]
        if non_empty:
            combined[col] = " / ".join(non_empty)

    return combined


def encode_raw_rows(sheet: SheetData, max_rows: int = 30) -> str:
    """Encode the first N rows as a raw markdown table without header assumptions.

    Used for LLM-based header detection: the LLM sees raw row data
    and decides which rows form the header.

    Includes merged region info as a critical hint for header detection.
    """
    parts: list[str] = []

    # * Sheet metadata
    parts.append(f"## Sheet: {sheet.name}")
    parts.append(f"Rows: {sheet.row_count}, Columns: {sheet.col_count}")

    # * Merged regions (critical for detecting multi-row headers)
    merged = format_merged_regions(sheet)
    if merged:
        parts.append("\n### Merged Regions")
        for m in merged:
            parts.append(f"- {m}")

    # * Raw table — all rows treated equally (no header/data distinction)
    capped_rows = min(max_rows, sheet.row_count)

    # ^ Determine active columns
    active_cols: list[int] = sorted(
        {c.col for c in sheet.cells if c.display_value is not None and c.row <= capped_rows}
    )
    if not active_cols:
        parts.append("\n(empty sheet)")
        return "\n".join(parts)

    # ^ Table header: column letters
    col_headers = ["Row"] + [get_column_letter(c) for c in active_cols]
    parts.append("")
    parts.append("| " + " | ".join(col_headers) + " |")
    parts.append("| " + " | ".join("---" for _ in col_headers) + " |")

    # ^ Pre-group cells by row for O(N) lookup instead of O(N*M)
    active_cols_set = set(active_cols)
    cells_by_row: dict[int, dict[int, str]] = {}
    for cell in sheet.cells:
        if cell.row <= capped_rows and cell.col in active_cols_set:
            cells_by_row.setdefault(cell.row, {})[cell.col] = format_cell_value(cell)

    # ^ Data rows
    for row_num in range(1, capped_rows + 1):
        row_values = cells_by_row.get(row_num, {})

        # ^ Skip completely empty rows
        if not any(row_values.get(c) for c in active_cols):
            continue

        row_parts = [str(row_num)]
        for col in active_cols:
            row_parts.append(row_values.get(col, ""))
        parts.append("| " + " | ".join(row_parts) + " |")

    return "\n".join(parts)


# * Merged cell formatting


def format_merged_regions(sheet: SheetData) -> list[str]:
    """Format merged regions as 'A1:C1 = "Header Text"' strings."""
    results: list[str] = []
    for mr_str in sheet.merged_ranges:
        # ^ Find the top-left cell value for this merge range
        match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", mr_str)
        if not match:
            results.append(mr_str)
            continue

        min_row = int(match.group(2))
        # ^ Find the min_col from the column letter
        min_col_letter = match.group(1)
        # ^ Find the origin cell value
        origin_cell = sheet.get_cell(min_row, _col_letter_to_num(min_col_letter))
        if origin_cell and origin_cell.display_value is not None:
            results.append(f'{mr_str} = "{origin_cell.display_value}"')
        else:
            results.append(mr_str)

    return results


# * Formula summary


def summarize_formulas(sheet: SheetData) -> list[str]:
    """Summarize formula cells, compressing repeated patterns.

    E.g., E3="=C3*D3", E4="=C4*D4", ... → E{3..100}="=C{r}*D{r}"
    """
    formula_cells: dict[int, list[CellData]] = {}
    for cell in sheet.cells:
        if cell.formula:
            formula_cells.setdefault(cell.col, []).append(cell)

    results: list[str] = []
    for col in sorted(formula_cells):
        cells = sorted(formula_cells[col], key=lambda c: c.row)
        col_letter = get_column_letter(col)

        if len(cells) <= 2:
            # ^ Few formulas: list individually
            for c in cells:
                results.append(f'{col_letter}{c.row}="{c.formula}"')
            continue

        # ^ Try to detect a pattern by replacing the row number
        first = cells[0]
        pattern = _generalize_formula(first.formula or "", first.row)
        matching = [c for c in cells if _generalize_formula(c.formula or "", c.row) == pattern]

        if len(matching) == len(cells):
            # ^ All cells share the same pattern
            start_row = cells[0].row
            end_row = cells[-1].row
            results.append(f'{col_letter}{{{start_row}..{end_row}}}="{pattern}"')
        else:
            # ^ Mixed patterns: list individually
            for c in cells:
                results.append(f'{col_letter}{c.row}="{c.formula}"')

    return results


def _generalize_formula(formula: str, row: int) -> str:
    """Replace concrete row numbers in cell references with {r} placeholder.

    Only replaces row numbers that appear as part of cell references (e.g., A12, BC3)
    to avoid corrupting numeric literals in formulas.
    """
    # ^ Match cell references: one or more uppercase letters followed by the row number
    # ^ Use word boundary after digits to avoid partial matches (e.g., 12 in 120)
    return re.sub(rf"([A-Z]+){row}\b", r"\g<1>{r}", formula)


# * Empty row/column detection


def find_empty_rows(sheet: SheetData) -> set[int]:
    """Find row numbers that are completely empty."""
    non_empty_rows: set[int] = set()
    for cell in sheet.cells:
        if cell.display_value is not None:
            non_empty_rows.add(cell.row)
    all_rows = set(range(1, sheet.row_count + 1))
    return all_rows - non_empty_rows


# * Column type summary


def _classify_number_format(nf: str) -> str | None:
    """Infer a semantic type from an Excel number format string.

    Returns "currency", "percentage", or "date" if the format matches
    well-known patterns; None otherwise.
    """
    # ^ Currency: contains $, EUR, GBP, JPY, or other currency symbols
    if re.search(r"[$€£¥₩]", nf) or re.search(r'"\$"', nf):
        return "currency"
    # ^ Percentage: contains literal %
    if "%" in nf:
        return "percentage"
    # ^ Date/time: contains y/m/d or h/m/s patterns (case-insensitive)
    if re.search(r"[yYmMdDhHsS]", nf) and re.search(r"[yYdDhHsS]", nf):
        return "date"
    return None


def summarize_column_types(
    sheet: SheetData,
    header_row: int | None = None,
) -> dict[int, str]:
    """Summarize the dominant data type per column.

    Returns {col: "int"/"float"/"str"/"date"/"bool"/"currency"/"percentage"/"mixed"}.

    When ``number_format`` is available on cells, it is used to refine
    numeric types into more specific semantic types such as "currency",
    "percentage", or "date".
    """
    col_types: dict[int, list[str]] = {}

    for cell in sheet.cells:
        if header_row and cell.row <= header_row:
            continue
        if cell.display_value is None:
            continue
        if cell.merge_origin is not None:
            continue

        # ^ Try number_format-based classification first
        nf_type: str | None = None
        if cell.number_format:
            nf_type = _classify_number_format(cell.number_format)

        val = cell.display_value
        if nf_type:
            t = nf_type
        elif isinstance(val, bool):
            t = "bool"
        elif isinstance(val, int):
            t = "int"
        elif isinstance(val, float):
            t = "float"
        elif isinstance(val, (datetime, date)):
            t = "date"
        else:
            t = "str"

        col_types.setdefault(cell.col, []).append(t)

    result: dict[int, str] = {}
    for col, types in col_types.items():
        counter = Counter(types)
        most_common, count = counter.most_common(1)[0]
        # ^ If dominant type covers >80% of values, use it; else "mixed"
        if count / len(types) >= 0.8:
            result[col] = most_common
        else:
            result[col] = "mixed"

    return result


# * Internal helpers


def _col_letter_to_num(letter: str) -> int:
    """Convert column letter (A, B, ..., Z, AA, ...) to 1-based number."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result
