"""CompressedEncoder: Markdown-format encoder with sampling and metadata.

Encodes sheet data as a markdown table with structural metadata
(column types, formula patterns, statistics). Supports full or sampled output.
"""

from openpyxl.utils import get_column_letter

from xlstruct.encoder._formatting import (
    build_column_headers,
    build_multi_row_headers,
    detect_header_row,
    find_empty_rows,
    format_cell_value,
    format_merged_regions,
    summarize_column_types,
    summarize_formulas,
)
from xlstruct.schemas.core import SheetData


class CompressedEncoder:
    """Encodes sheet data as a markdown table with structural metadata.

    When sample_size is set, only N data rows are included (head + tail sampling).
    Metadata (column types, formula patterns, statistics) is always included
    to give LLMs structural context even with sampled data.
    """

    def __init__(self, sample_size: int | None = None) -> None:
        self._sample_size = sample_size

    def encode(self, sheet: SheetData, header_rows: list[int] | None = None) -> str:
        parts: list[str] = []

        # * Sheet header
        parts.append(f'## Sheet: "{sheet.name}"')
        dim_info = f"{sheet.row_count} rows x {sheet.col_count} cols"
        parts.append(f"Dimensions: {sheet.dimensions} ({dim_info})")

        # * Merged regions
        merged = format_merged_regions(sheet)
        if merged:
            parts.append("\n### Merged Regions")
            for m in merged:
                parts.append(f"- {m}")

        # * Formula summary
        formulas = summarize_formulas(sheet)
        if formulas:
            parts.append("\n### Formulas")
            for f in formulas:
                parts.append(f"- {f}")

        # * Header: user-provided vs auto-detected
        if header_rows is not None:
            headers = build_multi_row_headers(sheet, header_rows)
            data_start = max(header_rows) + 1
        else:
            header_row = detect_header_row(sheet)
            headers = build_column_headers(sheet, header_row) if header_row else {}
            data_start = (header_row + 1) if header_row else 1

        # * Column types
        types = summarize_column_types(sheet, data_start - 1 if data_start > 1 else None)
        if types:
            active_cols = sorted(types.keys())
            type_parts = [
                f"{get_column_letter(c)} ({headers.get(c, '?')}): {types[c]}" for c in active_cols
            ]
            parts.append("\n### Column Types")
            parts.append(", ".join(type_parts))

        # * Build markdown table (full or sampled)
        empty_rows = find_empty_rows(sheet)
        table_text, total_data_rows, shown_data_rows = self._build_table(
            sheet, headers, data_start, empty_rows
        )

        parts.append("")
        if shown_data_rows < total_data_rows:
            parts.append(f"### Data (sample {shown_data_rows} of {total_data_rows} data rows)")
        parts.append(table_text)

        # * Stats
        parts.append(f"\nStats: rows={sheet.row_count}, data_rows={total_data_rows}")

        return "\n".join(parts)

    def _build_table(
        self,
        sheet: SheetData,
        headers: dict[int, str],
        data_start: int,
        empty_rows: set[int],
    ) -> tuple[str, int, int]:
        """Build markdown table, optionally sampling rows.

        Returns:
            (table_text, total_data_rows, shown_data_rows)
        """
        if not sheet.cells:
            return "(empty sheet)", 0, 0

        # ^ Determine columns to include (skip fully empty columns)
        all_cols: set[int] = set()
        for cell in sheet.cells:
            if cell.display_value is not None:
                all_cols.add(cell.col)
        if not all_cols:
            return "(empty sheet)", 0, 0

        cols = sorted(all_cols)

        # * Collect data rows
        data_rows: list[tuple[int, dict[int, str]]] = []
        for row_cells in sheet.iter_rows():
            row_num = row_cells[0].row
            if row_num < data_start:
                continue
            if row_num in empty_rows:
                continue

            row_values: dict[int, str] = {}
            for cell in row_cells:
                if cell.merge_origin is not None:
                    continue
                row_values[cell.col] = format_cell_value(cell)

            if not any(row_values.get(c) for c in cols):
                continue

            data_rows.append((row_num, row_values))

        # * Sample if needed (head + tail)
        if self._sample_size is not None and len(data_rows) > self._sample_size:
            head_count = self._sample_size // 2
            tail_count = self._sample_size - head_count
            sampled = data_rows[:head_count] + data_rows[-tail_count:]
        else:
            sampled = data_rows

        # * Table header row
        header_parts = ["| Row |"]
        separator_parts = ["|-----|"]
        for col in cols:
            col_letter = get_column_letter(col)
            label = headers.get(col, col_letter)
            header_parts.append(f" {col_letter} ({label}) |")
            separator_parts.append("------|")

        lines: list[str] = []
        lines.append("".join(header_parts))
        lines.append("".join(separator_parts))

        # * Data rows
        prev_row_num: int | None = None
        for row_num, row_values in sampled:
            # ^ Insert gap indicator when row numbers are non-contiguous
            if prev_row_num is not None and row_num > prev_row_num + 1:
                gap_line = "| ... |" + " ... |" * len(cols)
                lines.append(gap_line)

            row_parts = [f"| {row_num} |"]
            for col in cols:
                val = row_values.get(col, "")
                row_parts.append(f" {val} |")
            lines.append("".join(row_parts))
            prev_row_num = row_num

        return "\n".join(lines), len(data_rows), len(sampled)
