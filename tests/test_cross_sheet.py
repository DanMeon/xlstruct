"""Tests for cross-sheet extraction (extract_cross_sheet / extract_cross_sheet_sync)."""

import io
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest
from pydantic import BaseModel

from xlstruct.extractor import Extractor

# * Test schema


class QuarterlyRevenue(BaseModel):
    region: str
    quarter: str
    revenue: float


# * Fixtures


@pytest.fixture(autouse=True)
def _mock_instructor():
    """Mock build_instructor_client to avoid real API calls."""
    with patch(
        "xlstruct.extraction.engine.build_instructor_client",
        return_value=MagicMock(),
    ):
        yield


def _make_quarterly_xlsx(tmp_path) -> str:
    """Create an xlsx file with Q1 and Q2 sheets."""
    wb = openpyxl.Workbook()

    ws_q1 = wb.active
    ws_q1.title = "Q1"
    ws_q1["A1"] = "Region"
    ws_q1["B1"] = "Revenue"
    ws_q1["A2"] = "North"
    ws_q1["B2"] = 1000.0
    ws_q1["A3"] = "South"
    ws_q1["B3"] = 800.0

    ws_q2 = wb.create_sheet("Q2")
    ws_q2["A1"] = "Region"
    ws_q2["B1"] = "Revenue"
    ws_q2["A2"] = "North"
    ws_q2["B2"] = 1200.0
    ws_q2["A3"] = "South"
    ws_q2["B3"] = 950.0

    # ^ Extra sheet not used in cross-sheet extraction
    ws_q3 = wb.create_sheet("Q3")
    ws_q3["A1"] = "Region"
    ws_q3["B1"] = "Revenue"
    ws_q3["A2"] = "North"
    ws_q3["B2"] = 1100.0

    buf = io.BytesIO()
    wb.save(buf)
    path = tmp_path / "quarterly.xlsx"
    path.write_bytes(buf.getvalue())
    return str(path)


@pytest.fixture
def quarterly_xlsx(tmp_path) -> str:
    return _make_quarterly_xlsx(tmp_path)


# * Cross-sheet extraction tests


class TestExtractCrossSheet:
    async def test_two_sheet_extraction(self, quarterly_xlsx):
        """Two sheets are encoded and combined into a single LLM call."""
        extractor = Extractor()

        mock_results = [
            QuarterlyRevenue(region="North", quarter="Q1", revenue=1000.0),
            QuarterlyRevenue(region="South", quarter="Q1", revenue=800.0),
            QuarterlyRevenue(region="North", quarter="Q2", revenue=1200.0),
            QuarterlyRevenue(region="South", quarter="Q2", revenue=950.0),
        ]

        async def mock_extract(encoded: str, schema: Any, instructions: Any, **kw: Any) -> list:
            # ^ Verify both sheets appear in the combined encoding
            assert '## Sheet: "Q1"' in encoded
            assert '## Sheet: "Q2"' in encoded
            return mock_results

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            side_effect=mock_extract,
        ):
            result = await extractor.extract_cross_sheet(
                quarterly_xlsx,
                schema=QuarterlyRevenue,
                sheets=["Q1", "Q2"],
            )

        assert len(result) == 4
        assert result[0].region == "North"
        assert result[0].quarter == "Q1"
        assert result.report.mode == "direct"

    async def test_header_rows_as_list(self, quarterly_xlsx):
        """A single list[int] applies the same header rows to all sheets."""
        extractor = Extractor()

        async def mock_extract(encoded: str, schema: Any, instructions: Any, **kw: Any) -> list:
            return [QuarterlyRevenue(region="North", quarter="Q1", revenue=1000.0)]

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            side_effect=mock_extract,
        ):
            result = await extractor.extract_cross_sheet(
                quarterly_xlsx,
                schema=QuarterlyRevenue,
                sheets=["Q1", "Q2"],
                header_rows=[1],
            )

        assert len(result) == 1

    async def test_header_rows_as_dict(self, quarterly_xlsx):
        """Per-sheet header_rows via dict[str, list[int]]."""
        extractor = Extractor()

        async def mock_extract(encoded: str, schema: Any, instructions: Any, **kw: Any) -> list:
            return [QuarterlyRevenue(region="South", quarter="Q2", revenue=950.0)]

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            side_effect=mock_extract,
        ):
            result = await extractor.extract_cross_sheet(
                quarterly_xlsx,
                schema=QuarterlyRevenue,
                sheets=["Q1", "Q2"],
                header_rows={"Q1": [1], "Q2": [1]},
            )

        assert len(result) == 1
        assert result[0].quarter == "Q2"

    async def test_header_rows_dict_partial(self, quarterly_xlsx):
        """Dict header_rows with only some sheets specified falls back to None."""
        extractor = Extractor()

        async def mock_extract(encoded: str, schema: Any, instructions: Any, **kw: Any) -> list:
            return []

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            side_effect=mock_extract,
        ):
            # ^ Q2 not in dict — should use auto-detection (None)
            result = await extractor.extract_cross_sheet(
                quarterly_xlsx,
                schema=QuarterlyRevenue,
                sheets=["Q1", "Q2"],
                header_rows={"Q1": [1]},
            )

        assert len(result) == 0

    async def test_instructions_passed_through(self, quarterly_xlsx):
        """Instructions kwarg is forwarded to the engine."""
        extractor = Extractor()

        async def mock_extract(encoded: str, schema: Any, instructions: Any, **kw: Any) -> list:
            assert instructions == "Combine Q1 and Q2 data"
            return []

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            side_effect=mock_extract,
        ):
            await extractor.extract_cross_sheet(
                quarterly_xlsx,
                schema=QuarterlyRevenue,
                sheets=["Q1", "Q2"],
                instructions="Combine Q1 and Q2 data",
            )


# * Error cases


class TestExtractCrossSheetErrors:
    async def test_fewer_than_two_sheets(self, quarterly_xlsx):
        """Raises ValueError when fewer than 2 sheets are specified."""
        extractor = Extractor()

        with pytest.raises(ValueError, match="at least 2 sheets"):
            await extractor.extract_cross_sheet(
                quarterly_xlsx,
                schema=QuarterlyRevenue,
                sheets=["Q1"],
            )

    async def test_empty_sheets_list(self, quarterly_xlsx):
        """Raises ValueError for empty sheets list."""
        extractor = Extractor()

        with pytest.raises(ValueError, match="at least 2 sheets"):
            await extractor.extract_cross_sheet(
                quarterly_xlsx,
                schema=QuarterlyRevenue,
                sheets=[],
            )

    async def test_missing_sheet(self, quarterly_xlsx):
        """Raises ValueError when a requested sheet does not exist."""
        extractor = Extractor()

        with pytest.raises(ValueError, match="Sheets not found.*NonExistent"):
            await extractor.extract_cross_sheet(
                quarterly_xlsx,
                schema=QuarterlyRevenue,
                sheets=["Q1", "NonExistent"],
            )


# * Sync wrapper


class TestExtractCrossSheetSync:
    def test_sync_wrapper(self, quarterly_xlsx):
        """extract_cross_sheet_sync delegates to the async method."""
        extractor = Extractor()

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            new_callable=AsyncMock,
            return_value=[QuarterlyRevenue(region="North", quarter="Q1", revenue=1000.0)],
        ):
            result = extractor.extract_cross_sheet_sync(
                quarterly_xlsx,
                schema=QuarterlyRevenue,
                sheets=["Q1", "Q2"],
            )

        assert len(result) == 1
        assert result[0].region == "North"
        assert hasattr(result, "report")
