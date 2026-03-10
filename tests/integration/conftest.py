"""Shared fixtures for cloud storage integration tests.

Requires Docker Compose services running:
    cd tests/integration && docker compose up -d
"""

import io
import subprocess
import time
from pathlib import Path as PathLibPath

import openpyxl
import pytest

COMPOSE_DIR = PathLibPath(__file__).parent
COMPOSE_FILE = COMPOSE_DIR / "docker-compose.yml"

# * Emulator connection details
MINIO_ENDPOINT = "http://localhost:9100"
MINIO_KEY = "minioadmin"
MINIO_SECRET = "minioadmin"
MINIO_BUCKET = "test-bucket"

AZURITE_CONN_STR = (
    "DefaultEndpointsProtocol=http;"
    "AccountName=devstoreaccount1;"
    "AccountKey=Eby8vdM02xNOcqFlqUwJPLlmEtlCDXJ1OUzFT50uSRZ6IFsu"
    "Fq2UVErCz4I6tq/K1SZFPTOtr/KBHBeksoGMGw==;"
    "BlobEndpoint=http://127.0.0.1:10000/devstoreaccount1;"
)
# ^ Azurite well-known credentials (public, not real secrets)
# ^ https://learn.microsoft.com/en-us/azure/storage/common/storage-use-azurite#well-known-storage-account-and-key
AZURITE_ACCOUNT_NAME = "devstoreaccount1"
AZURITE_ACCOUNT_KEY = (
    "Eby8vdM02xNOcqFlqUwJPLlmEtlCDXJ1OUzFT50uSRZ6IFsu"
    "Fq2UVErCz4I6tq/K1SZFPTOtr/KBHBeksoGMGw=="
)
AZURITE_CONTAINER = "test-container"

FAKE_GCS_ENDPOINT = "http://localhost:4443"
FAKE_GCS_BUCKET = "test-bucket"


# * Sample xlsx data

@pytest.fixture(scope="session")
def sample_xlsx_bytes() -> bytes:
    """Minimal xlsx for upload/download testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws["A1"] = "Name"
    ws["B1"] = "Price"
    ws["A2"] = "Apple"
    ws["B2"] = 1.5
    ws["A3"] = "Banana"
    ws["B3"] = 0.75
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# * Docker Compose lifecycle

def _is_compose_running() -> bool:
    """Check if all compose services are healthy."""
    result = subprocess.run(
        ["docker", "compose", "-f", str(COMPOSE_FILE), "ps", "--format", "json"],
        capture_output=True,
        text=True,
    )
    return result.returncode == 0 and "running" in result.stdout.lower()


@pytest.fixture(scope="session", autouse=True)
def _ensure_compose_up():
    """Start Docker Compose services if not already running."""
    if _is_compose_running():
        yield
        return

    subprocess.run(
        ["docker", "compose", "-f", str(COMPOSE_FILE), "up", "-d", "--wait"],
        check=True,
        timeout=120,
    )
    # ^ Give services a moment to stabilize
    time.sleep(2)
    yield
    # ^ Don't tear down — user may want to inspect or re-run


# * Per-provider setup fixtures

@pytest.fixture(scope="session")
def s3_setup(sample_xlsx_bytes: bytes):
    """Create MinIO bucket and upload test file. Returns s3:// path."""
    import s3fs

    fs = s3fs.S3FileSystem(
        key=MINIO_KEY,
        secret=MINIO_SECRET,
        endpoint_url=MINIO_ENDPOINT,
    )
    if not fs.exists(MINIO_BUCKET):
        fs.mkdir(MINIO_BUCKET)
    s3_path = f"{MINIO_BUCKET}/test.xlsx"
    with fs.open(s3_path, "wb") as f:
        f.write(sample_xlsx_bytes)
    return f"s3://{s3_path}"


@pytest.fixture(scope="session")
def azure_setup(sample_xlsx_bytes: bytes):
    """Create Azurite container and upload test file. Returns az:// path."""
    from azure.storage.blob import BlobServiceClient

    client = BlobServiceClient.from_connection_string(AZURITE_CONN_STR)
    try:
        client.create_container(AZURITE_CONTAINER)
    except Exception:
        pass  # ^ Already exists
    blob_client = client.get_blob_client(AZURITE_CONTAINER, "test.xlsx")
    blob_client.upload_blob(sample_xlsx_bytes, overwrite=True)
    return f"az://{AZURITE_CONTAINER}/test.xlsx"


@pytest.fixture(scope="session")
def gcs_setup(sample_xlsx_bytes: bytes):
    """Create fake GCS bucket and upload test file. Returns gs:// path."""
    import requests

    # ^ fake-gcs-server: create bucket via HTTP API
    requests.post(f"{FAKE_GCS_ENDPOINT}/storage/v1/b", json={"name": FAKE_GCS_BUCKET})

    # ^ Upload file via resumable upload API
    upload_url = (
        f"{FAKE_GCS_ENDPOINT}/upload/storage/v1/b/{FAKE_GCS_BUCKET}/o"
        f"?uploadType=media&name=test.xlsx"
    )
    requests.post(upload_url, data=sample_xlsx_bytes, headers={"Content-Type": "application/octet-stream"})

    return f"gs://{FAKE_GCS_BUCKET}/test.xlsx"
