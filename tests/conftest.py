"""Shared test fixtures for XLStruct."""

import io

import openpyxl
import pytest

from xlstruct.schemas.core import CellData, SheetData


@pytest.fixture
def simple_sheet() -> SheetData:
    """A simple sheet with 5 data rows, no formulas, no merges."""
    cells = []
    # ^ Header row
    headers = {1: "Item", 2: "Description", 3: "Qty", 4: "Price"}
    for col, name in headers.items():
        cells.append(CellData(row=1, col=col, value=name, data_type="s"))
    # ^ Data rows
    data = [
        ("WDG-001", "Widget Alpha", 10, 25.0),
        ("WDG-002", "Widget Beta", 5, 42.5),
        ("WDG-003", "Widget Gamma", 20, 15.75),
        ("WDG-004", "Widget Delta", 8, 33.0),
        ("WDG-005", "Widget Epsilon", 15, 28.5),
    ]
    for i, (item, desc, qty, price) in enumerate(data, start=2):
        cells.append(CellData(row=i, col=1, value=item, data_type="s"))
        cells.append(CellData(row=i, col=2, value=desc, data_type="s"))
        cells.append(CellData(row=i, col=3, value=qty, data_type="n"))
        cells.append(CellData(row=i, col=4, value=price, data_type="n"))

    return SheetData(
        name="Inventory",
        dimensions="A1:D6",
        cells=cells,
        merged_ranges=[],
        row_count=6,
        col_count=4,
    )


@pytest.fixture
def merged_sheet() -> SheetData:
    """A sheet with merged header cells."""
    cells = [
        # ^ Merged title spanning A1:D1
        CellData(row=1, col=1, value="Invoice #2024-001", data_type="s",
                 is_merged=True, merge_range="A1:D1"),
        CellData(row=1, col=2, value=None, data_type="n",
                 is_merged=True, merge_range="A1:D1", merge_origin=(1, 1)),
        CellData(row=1, col=3, value=None, data_type="n",
                 is_merged=True, merge_range="A1:D1", merge_origin=(1, 1)),
        CellData(row=1, col=4, value=None, data_type="n",
                 is_merged=True, merge_range="A1:D1", merge_origin=(1, 1)),
        # ^ Header row
        CellData(row=2, col=1, value="Item", data_type="s"),
        CellData(row=2, col=2, value="Qty", data_type="s"),
        CellData(row=2, col=3, value="Price", data_type="s"),
        CellData(row=2, col=4, value="Total", data_type="s"),
        # ^ Data
        CellData(row=3, col=1, value="Widget", data_type="s"),
        CellData(row=3, col=2, value=10, data_type="n"),
        CellData(row=3, col=3, value=25.0, data_type="n"),
        CellData(row=3, col=4, value=250.0, data_type="n",
                 formula="=B3*C3", cached_value=250.0),
    ]
    return SheetData(
        name="Invoice",
        dimensions="A1:D3",
        cells=cells,
        merged_ranges=["A1:D1"],
        row_count=3,
        col_count=4,
    )


@pytest.fixture
def simple_xlsx_bytes() -> bytes:
    """Generate a real .xlsx file in memory for reader tests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # ^ Header
    ws["A1"] = "Name"
    ws["B1"] = "Price"
    ws["C1"] = "Stock"

    # ^ Data
    ws["A2"] = "Apple"
    ws["B2"] = 1.5
    ws["C2"] = 100

    ws["A3"] = "Banana"
    ws["B3"] = 0.75
    ws["C3"] = 200

    ws["A4"] = "Cherry"
    ws["B4"] = 3.0
    ws["C4"] = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


