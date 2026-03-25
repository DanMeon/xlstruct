"""Tests for Extractor.stream() and stream_sync()."""

import io
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest
from pydantic import BaseModel

from xlstruct.config import ExtractionConfig, ExtractionMode
from xlstruct.extractor import Extractor

# * Test schemas


class Product(BaseModel):
    name: str
    price: float
    stock: int


# * Fixtures


@pytest.fixture(autouse=True)
def _mock_instructor():
    """Mock build_instructor_client to avoid real API calls."""
    with patch(
        "xlstruct.extraction.engine.build_instructor_client",
        return_value=MagicMock(),
    ):
        yield


@pytest.fixture
def product_xlsx_file(tmp_path) -> str:
    """Real xlsx with product data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws["A1"] = "Name"
    ws["B1"] = "Price"
    ws["C1"] = "Stock"
    ws["A2"] = "Apple"
    ws["B2"] = 1.5
    ws["C2"] = 100
    ws["A3"] = "Banana"
    ws["B3"] = 0.75
    ws["C3"] = 200
    buf = io.BytesIO()
    wb.save(buf)
    path = tmp_path / "products.xlsx"
    path.write_bytes(buf.getvalue())
    return str(path)


@pytest.fixture
def large_xlsx_file(tmp_path) -> str:
    """Xlsx with enough rows to trigger chunking (> 100 rows)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws["A1"] = "Name"
    ws["B1"] = "Price"
    ws["C1"] = "Stock"
    for i in range(2, 153):  # ^ 151 data rows → triggers chunking
        ws[f"A{i}"] = f"Product{i}"
        ws[f"B{i}"] = float(i)
        ws[f"C{i}"] = i * 10
    buf = io.BytesIO()
    wb.save(buf)
    path = tmp_path / "large_products.xlsx"
    path.write_bytes(buf.getvalue())
    return str(path)


# * Tests


class TestStreamBasic:
    async def test_stream_yields_correct_records(self, product_xlsx_file):
        """stream() should yield all records from a small sheet."""
        expected = [
            Product(name="Apple", price=1.5, stock=100),
            Product(name="Banana", price=0.75, stock=200),
        ]

        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = expected
            results = [item async for item in extractor.stream(product_xlsx_file, Product)]

        assert len(results) == 2
        assert results[0].name == "Apple"
        assert results[1].name == "Banana"
        assert results[1].price == 0.75

    async def test_stream_single_chunk(self, product_xlsx_file):
        """Small sheet should not trigger chunking — single engine call."""
        expected = [Product(name="Apple", price=1.5, stock=100)]

        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = expected
            results = [item async for item in extractor.stream(product_xlsx_file, Product)]

        assert len(results) == 1
        mock_extract.assert_called_once()

    async def test_stream_with_instructions(self, product_xlsx_file):
        """stream() should pass instructions through to the engine."""
        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = []
            _ = [
                item
                async for item in extractor.stream(
                    product_xlsx_file, Product, instructions="Focus on fruits"
                )
            ]

        call_args = mock_extract.call_args
        assert call_args[0][2] == "Focus on fruits"

    async def test_stream_requires_schema_or_config(self, product_xlsx_file):
        """stream() should raise ValueError if neither schema nor config provided."""
        extractor = Extractor()
        with pytest.raises(ValueError, match="Either schema or extraction_config"):
            async for _ in extractor.stream(product_xlsx_file):
                pass


class TestStreamChunked:
    async def test_stream_multiple_chunks_yields_incrementally(self, large_xlsx_file):
        """Chunked extraction should call engine once per chunk."""
        chunk_results = [
            [Product(name="P1", price=1.0, stock=10)],
            [Product(name="P2", price=2.0, stock=20)],
        ]

        extractor = Extractor()
        call_count = 0

        async def mock_extract(encoded, schema, instructions=None):
            nonlocal call_count
            # ^ Return different results for each chunk call
            idx = min(call_count, len(chunk_results) - 1)
            result = chunk_results[idx]
            call_count += 1
            return result

        with patch.object(extractor._engine, "extract", side_effect=mock_extract):
            results = [item async for item in extractor.stream(large_xlsx_file, Product)]

        # ^ Should have multiple engine calls (one per chunk)
        assert call_count >= 2
        # ^ First chunk's record should appear
        assert results[0].name == "P1"

    async def test_stream_chunked_preserves_order(self, large_xlsx_file):
        """Records from earlier chunks should appear before later chunks."""
        call_idx = 0

        async def mock_extract(encoded, schema, instructions=None):
            nonlocal call_idx
            current = call_idx
            call_idx += 1
            return [Product(name=f"Chunk{current}", price=float(current), stock=current)]

        extractor = Extractor()
        with patch.object(extractor._engine, "extract", side_effect=mock_extract):
            results = [item async for item in extractor.stream(large_xlsx_file, Product)]

        # ^ Verify order: Chunk0 before Chunk1
        names = [r.name for r in results]
        assert names[0] == "Chunk0"
        assert names[1] == "Chunk1"


class TestStreamConfigured:
    async def test_stream_with_extraction_config_direct(self, product_xlsx_file):
        """stream() with ExtractionConfig in direct mode should yield records."""
        expected = [Product(name="Apple", price=1.5, stock=100)]
        config = ExtractionConfig(
            output_schema=Product,
            mode=ExtractionMode.DIRECT,
            header_rows=[1],
        )

        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = expected
            results = [
                item async for item in extractor.stream(product_xlsx_file, extraction_config=config)
            ]

        assert len(results) == 1
        assert results[0].name == "Apple"

    async def test_stream_with_codegen_mode(self, product_xlsx_file):
        """stream() with codegen mode should yield all results at once."""
        expected = [
            Product(name="Apple", price=1.5, stock=100),
            Product(name="Banana", price=0.75, stock=200),
        ]
        config = ExtractionConfig(
            output_schema=Product,
            mode=ExtractionMode.CODEGEN,
            header_rows=[1],
        )

        extractor = Extractor()
        with (
            patch.object(extractor, "_get_codegen", return_value=MagicMock()),
            patch.object(extractor, "_run_codegen", new_callable=AsyncMock) as mock_codegen,
        ):
            mock_codegen.return_value = expected
            results = [
                item async for item in extractor.stream(product_xlsx_file, extraction_config=config)
            ]

        assert len(results) == 2
        assert results[0].name == "Apple"
        assert results[1].name == "Banana"
        mock_codegen.assert_called_once()


class TestStreamSync:
    def test_stream_sync_returns_iterator(self, product_xlsx_file):
        """stream_sync() should return an iterator of records."""
        expected = [
            Product(name="Apple", price=1.5, stock=100),
            Product(name="Banana", price=0.75, stock=200),
        ]

        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = expected
            results = list(extractor.stream_sync(product_xlsx_file, Product))

        assert len(results) == 2
        assert results[0].name == "Apple"
        assert results[1].price == 0.75
