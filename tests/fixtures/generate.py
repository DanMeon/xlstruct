"""Generate test Excel fixtures for xlstruct.

Creates diverse .xlsx, .xlsm, and .xls files covering:
- Simple flat tables
- Merged cells (horizontal, vertical, multi-level)
- Formulas and calculations
- Multi-sheet workbooks
- Large datasets (200 rows)
- Complex real-world layouts (invoice forms, pivot-style, nested groups)

Usage:
    uv run python tests/fixtures/generate.py
"""

import io
import json
import random
import zipfile
from datetime import date, timedelta
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
import xlwt
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FIXTURES_DIR = Path(__file__).resolve().parent / "data"
FIXTURES_DIR.mkdir(parents=True, exist_ok=True)

EXPECTED_DIR = Path(__file__).resolve().parent / "expected"
EXPECTED_DIR.mkdir(parents=True, exist_ok=True)

random.seed(42)

# * Shared data pools
FIRST_NAMES = [
    "James",
    "Mary",
    "Robert",
    "Patricia",
    "John",
    "Jennifer",
    "Michael",
    "Linda",
    "David",
    "Elizabeth",
    "William",
    "Barbara",
    "Richard",
    "Susan",
    "Joseph",
    "Jessica",
    "Thomas",
    "Sarah",
    "Charles",
    "Karen",
    "Daniel",
    "Lisa",
    "Matthew",
    "Nancy",
    "Anthony",
    "Betty",
    "Mark",
    "Margaret",
    "Donald",
    "Sandra",
    "Steven",
    "Ashley",
    "Paul",
    "Dorothy",
    "Andrew",
    "Kimberly",
    "Joshua",
    "Emily",
    "Kenneth",
    "Donna",
]
LAST_NAMES = [
    "Smith",
    "Johnson",
    "Williams",
    "Brown",
    "Jones",
    "Garcia",
    "Miller",
    "Davis",
    "Rodriguez",
    "Martinez",
    "Hernandez",
    "Lopez",
    "Gonzalez",
    "Wilson",
    "Anderson",
    "Thomas",
    "Taylor",
    "Moore",
    "Jackson",
    "Martin",
    "Lee",
    "Perez",
    "Thompson",
    "White",
    "Harris",
    "Sanchez",
    "Clark",
    "Ramirez",
    "Lewis",
    "Robinson",
]
DEPARTMENTS = [
    "Engineering",
    "Sales",
    "Marketing",
    "HR",
    "Finance",
    "Operations",
    "Legal",
    "Support",
]
PRODUCTS = [
    "Laptop Pro 15",
    "Wireless Mouse",
    "USB-C Hub",
    "Monitor 27in",
    "Keyboard MX",
    "Webcam HD",
    "Headset Pro",
    "Docking Station",
    "External SSD 1TB",
    "Graphics Tablet",
    "Desk Lamp LED",
    "Ergonomic Chair",
    "Standing Desk",
    "Cable Organizer",
    "Mousepad XL",
    "Portable Charger",
    "Bluetooth Speaker",
    "HDMI Cable 6ft",
    "Screen Protector",
    "Laptop Stand",
    "Surge Protector",
    "Whiteboard 48in",
    "Marker Set 12pk",
    "Desk Shelf Riser",
    "Privacy Screen",
]
CATEGORIES = ["Electronics", "Peripherals", "Furniture", "Accessories", "Office Supplies"]
REGIONS = ["North America", "Europe", "Asia Pacific", "Latin America"]
COUNTRIES = {
    "North America": ["USA", "Canada", "Mexico"],
    "Europe": ["UK", "Germany", "France"],
    "Asia Pacific": ["Japan", "Australia", "Singapore"],
    "Latin America": ["Brazil", "Argentina", "Chile"],
}
PAYMENT_METHODS = ["Credit Card", "Debit Card", "PayPal", "Wire Transfer", "Cash"]
TXN_STATUSES = ["Completed", "Completed", "Completed", "Completed", "Pending", "Refunded", "Failed"]

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
BOLD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="center")


def _rand_name() -> str:
    return f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"


def _rand_email(name: str) -> str:
    parts = name.lower().split()
    return f"{parts[0]}.{parts[1]}@acmecorp.com"


def _rand_date(start: date, end: date) -> date:
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, delta))


def _style_header_row(ws, row: int, col_start: int, col_end: int):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data_cell(ws, row: int, col: int):
    ws.cell(row=row, column=col).border = THIN_BORDER


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def _save_expected(filename: str, records: list[dict]) -> None:
    """Save expected records as JSON for golden-file testing."""
    out_path = EXPECTED_DIR / filename
    out_path.write_text(json.dumps(records, indent=2, ensure_ascii=False))


# * Spreadsheet XML namespace
_SS_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _save_with_formula_cache(
    wb: openpyxl.Workbook,
    path: Path,
    cache: dict[str, dict[str, float | int | str]],
) -> None:
    """Save workbook and post-process xlsx XML to inject cached values for formula cells.

    Args:
        wb: openpyxl Workbook to save.
        path: Destination file path (.xlsx or .xlsm).
        cache: Mapping of sheet_name -> {cell_ref -> cached_value}.
               e.g. {"Invoice": {"F20": 8500.0, "F21": 6000.0}}
    """
    # ^ Save to an in-memory buffer first
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # ^ Parse workbook.xml to build sheet-name → internal XML path mapping
    with zipfile.ZipFile(buf, "r") as zin:
        names = zin.namelist()

        # ^ Read workbook.xml to get sheet rId → name mapping
        wb_xml_bytes = zin.read("xl/workbook.xml")
        # ^ Register namespaces to avoid ns0: prefixes on write
        ET.register_namespace("", _SS_NS)
        wb_tree = ET.fromstring(wb_xml_bytes)

        # ^ Build rId → sheet_name map from <sheet> elements
        rid_to_name: dict[str, str] = {}
        for sheet_el in wb_tree.iter(f"{{{_SS_NS}}}sheet"):
            r_id = sheet_el.get(f"{{{_REL_NS}}}id") or sheet_el.get("r:id") or ""
            s_name = sheet_el.get("name") or ""
            if r_id and s_name:
                rid_to_name[r_id] = s_name

        # ^ Read workbook.xml.rels to get rId → file path mapping
        rels_path = "xl/_rels/workbook.xml.rels"
        rid_to_file: dict[str, str] = {}
        if rels_path in names:
            rels_bytes = zin.read(rels_path)
            rels_tree = ET.fromstring(rels_bytes)
            for rel_el in rels_tree:
                r_id = rel_el.get("Id") or ""
                target = rel_el.get("Target") or ""
                if r_id and target:
                    # ^ Targets may be absolute (start with /) or relative to xl/
                    if target.startswith("/"):
                        # ^ Absolute path: strip leading slash to get zip entry path
                        file_path = target.lstrip("/")
                    else:
                        # ^ Relative to xl/ directory
                        file_path = f"xl/{target}"
                    rid_to_file[r_id] = file_path

        # ^ Build sheet_name → xml_file_path map
        name_to_file: dict[str, str] = {
            rid_to_name[rid]: rid_to_file[rid] for rid in rid_to_name if rid in rid_to_file
        }

        # ^ Collect all namespace declarations from each sheet XML we'll modify
        # ^ so we can re-register them and avoid dropping prefixes on serialisation
        modified: dict[str, bytes] = {}
        for sheet_name, cell_cache in cache.items():
            xml_path = name_to_file.get(sheet_name)
            if xml_path is None or xml_path not in names:
                continue
            sheet_bytes = zin.read(xml_path)

            # ^ Collect all namespace prefixes from the raw XML declaration
            # ^ and register them so ET doesn't mangle them
            import re as _re

            for prefix, uri in _re.findall(r'xmlns:?(\w*)="([^"]+)"', sheet_bytes.decode("utf-8")):
                if prefix:
                    ET.register_namespace(prefix, uri)
                else:
                    ET.register_namespace("", uri)

            sheet_tree = ET.fromstring(sheet_bytes)

            # ^ Build a map of cell ref → <c> element for fast lookup
            ref_to_cell: dict[str, ET.Element] = {}
            for c_el in sheet_tree.iter(f"{{{_SS_NS}}}c"):
                ref = c_el.get("r")
                if ref:
                    ref_to_cell[ref] = c_el

            # ^ Inject <v> tags for formula cells in the cache
            for cell_ref, cached_value in cell_cache.items():
                c_el = ref_to_cell.get(cell_ref)
                if c_el is None:
                    continue
                # ^ Only inject when cell has a formula (<f> child) but no cached value (<v>)
                f_el = c_el.find(f"{{{_SS_NS}}}f")
                if f_el is None:
                    continue
                v_el = c_el.find(f"{{{_SS_NS}}}v")
                if v_el is None:
                    v_el = ET.SubElement(c_el, f"{{{_SS_NS}}}v")
                v_el.text = str(cached_value)

            modified[xml_path] = ET.tostring(sheet_tree, encoding="unicode").encode("utf-8")

        # ^ Rebuild the zip with modified sheet XMLs
        out_buf = io.BytesIO()
        buf.seek(0)
        with (
            zipfile.ZipFile(buf, "r") as zin2,
            zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout,
        ):
            for item in zin2.infolist():
                if item.filename in modified:
                    zout.writestr(item, modified[item.filename])
                else:
                    zout.writestr(item, zin2.read(item.filename))

    path.write_bytes(out_buf.getvalue())


# * ──────────────────────────────────────────────────────────────────────
# * .xlsx files
# * ──────────────────────────────────────────────────────────────────────


def create_simple_employee_directory() -> list[dict]:
    """30-row flat employee table with mixed data types."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = [
        "Employee ID",
        "Full Name",
        "Department",
        "Email",
        "Hire Date",
        "Annual Salary",
        "Active",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, 1, len(headers))

    records: list[dict] = []
    for i in range(2, 32):
        name = _rand_name()
        dept = random.choice(DEPARTMENTS)
        hire = _rand_date(date(2018, 1, 1), date(2024, 12, 31))
        salary = round(random.uniform(45000, 180000), 2)
        active = random.random() > 0.15

        ws.cell(row=i, column=1, value=f"EMP-{1000 + i - 1}")
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=dept)
        ws.cell(row=i, column=4, value=_rand_email(name))
        ws.cell(row=i, column=5, value=hire)
        ws.cell(row=i, column=5).number_format = "YYYY-MM-DD"
        ws.cell(row=i, column=6, value=salary)
        ws.cell(row=i, column=6).number_format = "#,##0.00"
        ws.cell(row=i, column=7, value=active)

        for col in range(1, len(headers) + 1):
            _style_data_cell(ws, i, col)

        records.append(
            {
                "employee_id": f"EMP-{1000 + i - 1}",
                "full_name": name,
                "department": dept,
                "email": _rand_email(name),
                "hire_date": hire.isoformat(),
                "annual_salary": salary,
                "active": active,
            }
        )

    _auto_width(ws)
    wb.save(FIXTURES_DIR / "simple_employee_directory.xlsx")
    print("  Created simple_employee_directory.xlsx (30 rows)")
    return records


def create_merged_financial_report() -> list[dict]:
    """Financial report with 2-level merged headers, category groups, and formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annual Report"

    # * Title row
    ws.merge_cells("A1:M1")
    ws["A1"] = "Annual Financial Report — FY 2024"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    # * Quarter headers (row 3) — merged across 3 month columns each
    quarters = ["Q1", "Q2", "Q3", "Q4"]
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    ws.cell(row=3, column=1, value="Category")
    ws.cell(row=4, column=1, value="")

    for qi, q in enumerate(quarters):
        start_col = 2 + qi * 3
        end_col = start_col + 2
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
        cell = ws.cell(row=3, column=start_col, value=q)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # * Month sub-headers (row 4)
    ws.cell(row=3, column=1).font = BOLD_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=1).font = HEADER_FONT
    ws.merge_cells("A3:A4")
    ws.cell(row=3, column=1, value="Category")

    # ^ "Total" column
    ws.cell(row=3, column=14, value="FY Total")
    ws.cell(row=3, column=14).font = HEADER_FONT
    ws.cell(row=3, column=14).fill = HEADER_FILL
    ws.cell(row=3, column=14).alignment = CENTER
    ws.merge_cells(start_row=3, start_column=14, end_row=4, end_column=14)

    for mi, m in enumerate(months):
        col = 2 + mi
        cell = ws.cell(row=4, column=col, value=m)
        cell.font = Font(bold=True, size=10)
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER

    # * Data section
    revenue_items = {
        "Product Sales": [120, 135, 142, 158, 165, 170, 180, 175, 190, 210, 225, 240],
        "Service Revenue": [45, 48, 50, 52, 55, 58, 60, 62, 65, 68, 70, 75],
        "Licensing Fees": [30, 30, 32, 32, 35, 35, 35, 38, 38, 40, 40, 42],
    }
    expense_items = {
        "Salaries & Benefits": [85, 85, 88, 88, 90, 90, 92, 92, 95, 95, 98, 98],
        "Marketing & Ads": [25, 30, 28, 35, 40, 38, 42, 45, 50, 55, 60, 65],
        "R&D Costs": [40, 42, 45, 43, 48, 50, 52, 50, 55, 58, 60, 62],
        "Cloud Infrastructure": [15, 16, 17, 18, 19, 20, 22, 23, 25, 26, 28, 30],
        "Office & Operations": [10, 10, 10, 12, 12, 12, 13, 13, 14, 14, 15, 15],
        "Travel & Events": [8, 5, 12, 6, 10, 15, 8, 5, 18, 12, 8, 20],
    }

    row = 6
    section_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # * Revenue section — merged "REVENUE" label vertically
    ws.cell(row=row, column=1, value="REVENUE").font = SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    records: list[dict] = []
    # ^ formula_cache tracks {cell_ref: computed_value} for post-processing
    formula_cache: dict[str, float | int | str] = {}

    # ^ col_data accumulates raw values per column (col index 2..14) for subtotal computation
    col_values: dict[int, list[float]] = {c: [] for c in range(2, 15)}

    rev_data_start = row
    for item_name, values in revenue_items.items():
        ws.cell(row=row, column=1, value=item_name)
        for mi, v in enumerate(values):
            # ^ Values in thousands
            ws.cell(row=row, column=2 + mi, value=v * 1000)
            ws.cell(row=row, column=2 + mi).number_format = "#,##0"
            col_values[2 + mi].append(v * 1000)
        # ^ FY Total formula
        col_start = get_column_letter(2)
        col_end = get_column_letter(13)
        fy_total = float(sum(v * 1000 for v in values))
        ws.cell(row=row, column=14, value=f"=SUM({col_start}{row}:{col_end}{row})")
        ws.cell(row=row, column=14).number_format = "#,##0"
        formula_cache[f"N{row}"] = fy_total
        col_values[14].append(fy_total)
        records.append(
            {
                "section": "Revenue",
                "category": item_name,
                "fy_total": fy_total,
            }
        )
        row += 1

    # * Revenue subtotal
    ws.cell(row=row, column=1, value="Total Revenue")
    ws.cell(row=row, column=1).font = BOLD_FONT
    # ^ Compute column-wise sums for revenue rows
    rev_col_sums: dict[int, float] = {}
    for col in range(2, 15):
        cl = get_column_letter(col)
        col_sum = sum(col_values[col])
        rev_col_sums[col] = col_sum
        ws.cell(row=row, column=col, value=f"=SUM({cl}{rev_data_start}:{cl}{row - 1})")
        ws.cell(row=row, column=col).number_format = "#,##0"
        ws.cell(row=row, column=col).font = BOLD_FONT
        formula_cache[f"{cl}{row}"] = col_sum
    rev_total_row = row
    row += 2

    # * Reset col_values for expense section
    col_values = {c: [] for c in range(2, 15)}

    # * Expenses section
    ws.cell(row=row, column=1, value="EXPENSES").font = SECTION_FONT
    exp_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.cell(row=row, column=1).fill = exp_fill
    row += 1

    exp_data_start = row
    for item_name, values in expense_items.items():
        ws.cell(row=row, column=1, value=item_name)
        for mi, v in enumerate(values):
            ws.cell(row=row, column=2 + mi, value=v * 1000)
            ws.cell(row=row, column=2 + mi).number_format = "#,##0"
            col_values[2 + mi].append(v * 1000)
        col_s = get_column_letter(2)
        col_e = get_column_letter(13)
        fy_total = float(sum(v * 1000 for v in values))
        ws.cell(row=row, column=14, value=f"=SUM({col_s}{row}:{col_e}{row})")
        ws.cell(row=row, column=14).number_format = "#,##0"
        formula_cache[f"N{row}"] = fy_total
        col_values[14].append(fy_total)
        records.append(
            {
                "section": "Expenses",
                "category": item_name,
                "fy_total": fy_total,
            }
        )
        row += 1

    # * Expense subtotal
    ws.cell(row=row, column=1, value="Total Expenses")
    ws.cell(row=row, column=1).font = BOLD_FONT
    exp_col_sums: dict[int, float] = {}
    for col in range(2, 15):
        cl = get_column_letter(col)
        col_sum = sum(col_values[col])
        exp_col_sums[col] = col_sum
        ws.cell(row=row, column=col, value=f"=SUM({cl}{exp_data_start}:{cl}{row - 1})")
        ws.cell(row=row, column=col).number_format = "#,##0"
        ws.cell(row=row, column=col).font = BOLD_FONT
        formula_cache[f"{cl}{row}"] = col_sum
    exp_total_row = row
    row += 2

    # * Net Income
    net_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    ws.cell(row=row, column=1, value="NET INCOME")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = net_fill
    for col in range(2, 15):
        cl = get_column_letter(col)
        net_val = rev_col_sums.get(col, 0.0) - exp_col_sums.get(col, 0.0)
        ws.cell(
            row=row,
            column=col,
            value=f"={cl}{rev_total_row}-{cl}{exp_total_row}",
        )
        ws.cell(row=row, column=col).number_format = "#,##0"
        ws.cell(row=row, column=col).font = Font(bold=True, size=12)
        ws.cell(row=row, column=col).fill = net_fill
        formula_cache[f"{cl}{row}"] = net_val

    _auto_width(ws)
    ws.column_dimensions["A"].width = 22
    _save_with_formula_cache(
        wb, FIXTURES_DIR / "merged_financial_report.xlsx", {"Annual Report": formula_cache}
    )
    print("  Created merged_financial_report.xlsx (~25 rows, merged headers + formulas)")
    return records


def create_large_transaction_log() -> list[dict]:
    """200-row transaction log with summary sheet."""
    wb = openpyxl.Workbook()

    # * Sheet 1: Transactions
    ws1 = wb.active
    ws1.title = "Transactions"

    headers = [
        "Txn ID",
        "Date",
        "Time",
        "Customer ID",
        "Customer Name",
        "Category",
        "Description",
        "Amount",
        "Payment Method",
        "Status",
    ]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header_row(ws1, 1, 1, len(headers))

    records: list[dict] = []
    base_date = date(2024, 1, 1)
    # ^ Accumulators for computing formula caches for Summary/Monthly sheets
    cat_count: dict[str, int] = {c: 0 for c in CATEGORIES}
    cat_amount: dict[str, float] = {c: 0.0 for c in CATEGORIES}
    month_count: dict[int, int] = {m: 0 for m in range(1, 13)}
    month_amount: dict[int, float] = {m: 0.0 for m in range(1, 13)}
    txn_dates: list[date] = []

    for i in range(2, 202):
        txn_date = _rand_date(base_date, date(2024, 12, 31))
        hour = random.randint(8, 21)
        minute = random.randint(0, 59)
        cat = random.choice(CATEGORIES)
        prod = random.choice(PRODUCTS)
        amt = round(random.uniform(9.99, 4999.99), 2)
        payment = random.choice(PAYMENT_METHODS)
        status = random.choice(TXN_STATUSES)
        customer_name = _rand_name()

        ws1.cell(row=i, column=1, value=f"TXN-{10000 + i - 1}")
        ws1.cell(row=i, column=2, value=txn_date)
        ws1.cell(row=i, column=2).number_format = "YYYY-MM-DD"
        ws1.cell(row=i, column=3, value=f"{hour:02d}:{minute:02d}")
        ws1.cell(row=i, column=4, value=f"CUST-{random.randint(1000, 9999)}")
        ws1.cell(row=i, column=5, value=customer_name)
        ws1.cell(row=i, column=6, value=cat)
        ws1.cell(row=i, column=7, value=prod)
        ws1.cell(row=i, column=8, value=amt)
        ws1.cell(row=i, column=8).number_format = "#,##0.00"
        ws1.cell(row=i, column=9, value=payment)
        ws1.cell(row=i, column=10, value=status)

        # ^ Track per-category and per-month aggregates for cache injection
        cat_count[cat] += 1
        cat_amount[cat] += amt
        month_count[txn_date.month] += 1
        month_amount[txn_date.month] += amt
        txn_dates.append(txn_date)

        records.append(
            {
                "txn_id": f"TXN-{10000 + i - 1}",
                "date": txn_date.isoformat(),
                "customer_name": customer_name,
                "category": cat,
                "description": prod,
                "amount": amt,
                "payment_method": payment,
                "status": status,
            }
        )

    _auto_width(ws1)

    # * Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Transaction Summary"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:D1")

    ws2["A3"] = "Category"
    ws2["B3"] = "Transaction Count"
    ws2["C3"] = "Total Amount"
    ws2["D3"] = "Average Amount"
    _style_header_row(ws2, 3, 1, 4)

    summary_cache: dict[str, float | int | str] = {}
    for ci, cat in enumerate(CATEGORIES):
        row = 4 + ci
        count = cat_count[cat]
        total_amt = round(cat_amount[cat], 2)
        avg_amt = round(total_amt / count, 2) if count > 0 else 0.0
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=2, value=f'=COUNTIF(Transactions!F:F,"{cat}")')
        ws2.cell(row=row, column=3, value=f'=SUMIF(Transactions!F:F,"{cat}",Transactions!H:H)')
        ws2.cell(row=row, column=3).number_format = "#,##0.00"
        ws2.cell(row=row, column=4, value=f"=IF(B{row}>0,C{row}/B{row},0)")
        ws2.cell(row=row, column=4).number_format = "#,##0.00"
        summary_cache[f"B{row}"] = count
        summary_cache[f"C{row}"] = total_amt
        summary_cache[f"D{row}"] = avg_amt

    total_row = 4 + len(CATEGORIES) + 1
    ws2.cell(row=total_row, column=1, value="GRAND TOTAL")
    ws2.cell(row=total_row, column=1).font = BOLD_FONT
    ws2.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row - 2})")
    ws2.cell(row=total_row, column=2).font = BOLD_FONT
    ws2.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row - 2})")
    ws2.cell(row=total_row, column=3).font = BOLD_FONT
    ws2.cell(row=total_row, column=3).number_format = "#,##0.00"
    summary_cache[f"B{total_row}"] = sum(cat_count.values())
    summary_cache[f"C{total_row}"] = round(sum(cat_amount.values()), 2)

    # * Sheet 3: Monthly Breakdown
    ws3 = wb.create_sheet("Monthly")
    ws3["A1"] = "Monthly Revenue Breakdown"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:D1")

    ws3["A3"] = "Month"
    ws3["B3"] = "Transactions"
    ws3["C3"] = "Revenue"
    _style_header_row(ws3, 3, 1, 3)

    month_names = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]
    monthly_cache: dict[str, float | int | str] = {}
    for mi, mname in enumerate(month_names):
        row = 4 + mi
        m_num = mi + 1
        ws3.cell(row=row, column=1, value=mname)
        # ^ SUMPRODUCT to count transactions by month
        ws3.cell(
            row=row,
            column=2,
            value=f"=SUMPRODUCT((MONTH(Transactions!B2:B201)={m_num})*1)",
        )
        ws3.cell(
            row=row,
            column=3,
            value=f"=SUMPRODUCT((MONTH(Transactions!B2:B201)={m_num})*Transactions!H2:H201)",
        )
        ws3.cell(row=row, column=3).number_format = "#,##0.00"
        monthly_cache[f"B{row}"] = month_count[m_num]
        monthly_cache[f"C{row}"] = round(month_amount[m_num], 2)

    _auto_width(ws2)
    _auto_width(ws3)
    _save_with_formula_cache(
        wb,
        FIXTURES_DIR / "large_transaction_log.xlsx",
        {"Summary": summary_cache, "Monthly": monthly_cache},
    )
    print("  Created large_transaction_log.xlsx (200 rows + 2 summary sheets)")
    return records


# * ──────────────────────────────────────────────────────────────────────
# * .xlsm files
# * ──────────────────────────────────────────────────────────────────────


def create_formula_budget_tracker() -> list[dict]:
    """Budget tracker with extensive formulas: SUM, IF, percentage."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget 2024"

    # * Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "Department Budget Tracker — FY 2024 (in USD)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    headers = ["Department", "Annual Budget"]
    month_abbrs = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]
    headers += [f"{m} Actual" for m in month_abbrs]
    headers += ["YTD Actual", "Variance", "% Used"]

    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header_row(ws, 3, 1, len(headers))

    budgets = {
        "Engineering": 2400000,
        "Sales": 1800000,
        "Marketing": 1200000,
        "HR": 600000,
        "Finance": 500000,
        "Operations": 900000,
        "Legal": 400000,
        "Support": 700000,
        "Executive": 350000,
        "IT Infrastructure": 800000,
        "Product": 1100000,
        "Data Science": 950000,
    }

    records: list[dict] = []
    formula_cache: dict[str, float | int | str] = {}
    # ^ Accumulate per-column sums for the totals row
    col_totals: dict[int, float] = {c: 0.0 for c in range(2, 17)}

    row = 4
    for dept, annual_budget in budgets.items():
        ws.cell(row=row, column=1, value=dept)
        ws.cell(row=row, column=2, value=annual_budget)
        ws.cell(row=row, column=2).number_format = "#,##0"
        col_totals[2] += annual_budget

        monthly_budget = annual_budget / 12
        monthly_actuals: list[float] = []
        for mi in range(12):
            # ^ Simulate actual spend: budget ± 20% random variance
            actual = round(monthly_budget * random.uniform(0.75, 1.25))
            ws.cell(row=row, column=3 + mi, value=actual)
            ws.cell(row=row, column=3 + mi).number_format = "#,##0"
            monthly_actuals.append(actual)
            col_totals[3 + mi] += actual

        ytd_actual = sum(monthly_actuals)
        variance = annual_budget - ytd_actual
        percent_used = ytd_actual / annual_budget if annual_budget > 0 else 0.0

        # ^ YTD Actual = SUM of monthly actuals
        ytd_col = 15
        ws.cell(row=row, column=ytd_col, value=f"=SUM(C{row}:N{row})")
        ws.cell(row=row, column=ytd_col).number_format = "#,##0"
        formula_cache[f"O{row}"] = ytd_actual
        col_totals[15] += ytd_actual

        # ^ Variance = Budget - YTD
        ws.cell(row=row, column=16, value=f"=B{row}-O{row}")
        ws.cell(row=row, column=16).number_format = "#,##0"
        formula_cache[f"P{row}"] = variance
        col_totals[16] += variance

        # ^ % Used = YTD / Budget
        ws.cell(row=row, column=17, value=f"=IF(B{row}>0,O{row}/B{row},0)")
        ws.cell(row=row, column=17).number_format = "0.0%"
        formula_cache[f"Q{row}"] = percent_used

        records.append(
            {
                "department": dept,
                "annual_budget": float(annual_budget),
                "ytd_actual": ytd_actual,
                "variance": variance,
                "percent_used": round(percent_used, 10),
            }
        )

        row += 1

    # * Totals row
    total_row = row
    ws.cell(row=total_row, column=1, value="COMPANY TOTAL")
    ws.cell(row=total_row, column=1).font = BOLD_FONT
    total_budget = col_totals[2]
    total_ytd = col_totals[15]
    for col in range(2, 18):
        cl = get_column_letter(col)
        if col == 17:
            total_pct = total_ytd / total_budget if total_budget > 0 else 0.0
            ws.cell(
                row=total_row, column=col, value=f"=IF(B{total_row}>0,O{total_row}/B{total_row},0)"
            )
            ws.cell(row=total_row, column=col).number_format = "0.0%"
            formula_cache[f"Q{total_row}"] = total_pct
        else:
            col_sum = col_totals.get(col, 0.0)
            ws.cell(row=total_row, column=col, value=f"=SUM({cl}4:{cl}{total_row - 1})")
            ws.cell(row=total_row, column=col).number_format = "#,##0"
            formula_cache[f"{cl}{total_row}"] = col_sum
        ws.cell(row=total_row, column=col).font = BOLD_FONT

    # * Conditional formatting note row
    row = total_row + 2
    ws.cell(row=row, column=1, value="Status Legend:")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row + 1, column=1, value="Under Budget: % Used < 90%")
    ws.cell(row=row + 2, column=1, value="On Track: 90% <= % Used <= 105%")
    ws.cell(row=row + 3, column=1, value="Over Budget: % Used > 105%")

    _auto_width(ws)
    ws.column_dimensions["A"].width = 20
    # ^ Save as .xlsm (macro-enabled)
    _save_with_formula_cache(
        wb, FIXTURES_DIR / "formula_budget_tracker.xlsm", {"Budget 2024": formula_cache}
    )
    print("  Created formula_budget_tracker.xlsm (12 departments, 12 months + formulas)")
    return records


def create_multi_level_header_sales() -> list[dict]:
    """Sales data with 3-level merged headers: Region → Country → Metric."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Regional Sales"

    # * Title
    ws.merge_cells("A1:M1")
    ws["A1"] = "Global Sales Report — 2024"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    # * Level 1: Region headers (row 3)
    ws.cell(row=3, column=1, value="Product")
    ws.merge_cells("A3:A5")
    ws.cell(row=3, column=1).font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=1).alignment = CENTER

    region_list = ["North America", "Europe", "Asia Pacific", "Latin America"]
    col = 2
    for region in region_list:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
        cell = ws.cell(row=3, column=col, value=region)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

        # * Level 2: Metric sub-headers (row 4) — Revenue | Units | Margin %
        metrics = ["Revenue", "Units", "Margin %"]
        for mi, metric in enumerate(metrics):
            c = ws.cell(row=4, column=col + mi, value=metric)
            c.font = Font(bold=True, size=10)
            c.fill = SUBHEADER_FILL
            c.alignment = CENTER

        col += 3

    # * Data rows: products grouped by category with vertical merges
    product_groups = {
        "Electronics": ["Laptop Pro 15", "Monitor 27in", "Webcam HD", "Headset Pro"],
        "Peripherals": ["Wireless Mouse", "Keyboard MX", "USB-C Hub", "Graphics Tablet"],
        "Furniture": ["Ergonomic Chair", "Standing Desk", "Desk Lamp LED", "Desk Shelf Riser"],
        "Accessories": [
            "Portable Charger",
            "Bluetooth Speaker",
            "Laptop Stand",
            "Cable Organizer",
            "Mousepad XL",
            "Screen Protector",
        ],
        "Office Supplies": [
            "Whiteboard 48in",
            "Marker Set 12pk",
            "Surge Protector",
            "HDMI Cable 6ft",
            "Privacy Screen",
        ],
    }

    records: list[dict] = []
    formula_cache: dict[str, float | int | str] = {}
    # ^ Accumulate per-column values for the SUBTOTAL formula (col 2..13)
    # ^ Category rows have no numeric values, so we track data-row values only
    col_data_values: dict[int, list[float]] = {c: [] for c in range(2, 14)}

    row = 5
    cat_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for cat_name, products in product_groups.items():
        # ^ Category separator row
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        ws.cell(row=row, column=1, value=cat_name)
        ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
        ws.cell(row=row, column=1).fill = cat_fill

        for prod in products:
            row += 1
            ws.cell(row=row, column=1, value=prod)
            col = 2
            for region in region_list:
                revenue = round(random.uniform(10000, 500000), 2)
                units = random.randint(50, 5000)
                margin = round(random.uniform(0.08, 0.45), 2)
                ws.cell(row=row, column=col, value=revenue)
                ws.cell(row=row, column=col).number_format = "#,##0.00"
                ws.cell(row=row, column=col + 1, value=units)
                ws.cell(row=row, column=col + 2, value=margin)
                ws.cell(row=row, column=col + 2).number_format = "0.0%"
                col_data_values[col].append(revenue)
                col_data_values[col + 1].append(float(units))
                col_data_values[col + 2].append(margin)
                records.append(
                    {
                        "category": cat_name,
                        "product": prod,
                        "region": region,
                        "revenue": revenue,
                        "units": units,
                        "margin_pct": margin,
                    }
                )
                col += 3

    # * Region totals
    row += 2
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = BOLD_FONT
    for c in range(2, 14):
        cl = get_column_letter(c)
        # ^ SUBTOTAL(9,...) = SUM when no rows are hidden
        col_sum = sum(col_data_values[c])
        ws.cell(row=row, column=c, value=f"=SUBTOTAL(9,{cl}6:{cl}{row - 2})")
        ws.cell(row=row, column=c).font = BOLD_FONT
        ws.cell(row=row, column=c).number_format = "#,##0.00"
        formula_cache[f"{cl}{row}"] = col_sum

    _auto_width(ws)
    _save_with_formula_cache(
        wb, FIXTURES_DIR / "multi_level_header_sales.xlsm", {"Regional Sales": formula_cache}
    )
    print("  Created multi_level_header_sales.xlsm (3-level merged headers, ~35 products)")
    return records


def create_complex_invoice_form() -> list[dict]:
    """Invoice with form-style header (merged cells) + line items + formula totals."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # * Company header — large merged area
    ws.merge_cells("A1:F2")
    ws["A1"] = "ACME CORPORATION"
    ws["A1"].font = Font(bold=True, size=20)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = "123 Business Ave, Suite 500, San Francisco, CA 94102"
    ws["A3"].alignment = CENTER

    ws.merge_cells("A4:F4")
    ws["A4"] = "Tel: (415) 555-0100 | Email: billing@acmecorp.com"
    ws["A4"].alignment = CENTER

    # * "INVOICE" label
    ws.merge_cells("A6:F6")
    ws["A6"] = "INVOICE"
    ws["A6"].font = Font(bold=True, size=16, color="4472C4")
    ws["A6"].alignment = CENTER

    # * Invoice metadata — key-value pairs in merged cells
    meta = [
        ("Invoice Number:", "INV-2024-00847"),
        ("Invoice Date:", date(2024, 11, 15)),
        ("Due Date:", date(2024, 12, 15)),
        ("Payment Terms:", "Net 30"),
    ]
    for mi, (label, val) in enumerate(meta):
        row = 8 + mi
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3, value=val)
        if isinstance(val, date):
            ws.cell(row=row, column=3).number_format = "YYYY-MM-DD"

    # * Bill To section
    ws.merge_cells("A13:B13")
    ws["A13"] = "Bill To:"
    ws["A13"].font = BOLD_FONT

    ws.merge_cells("A14:D14")
    ws["A14"] = "TechStart Inc."
    ws.merge_cells("A15:D15")
    ws["A15"] = "456 Innovation Drive"
    ws.merge_cells("A16:D16")
    ws["A16"] = "Austin, TX 78701"
    ws.merge_cells("A17:D17")
    ws["A17"] = "Contact: Sarah Johnson | sarah@techstart.io"

    # * Line items table
    item_row = 19
    item_headers = ["#", "Description", "Qty", "Unit Price", "Tax Rate", "Line Total"]
    for col, h in enumerate(item_headers, 1):
        ws.cell(row=item_row, column=col, value=h)
    _style_header_row(ws, item_row, 1, 6)

    line_items = [
        ("Website Redesign — UX Research Phase", 1, 8500.00, 0.0),
        ("Website Redesign — UI Design (40 hrs)", 40, 150.00, 0.0),
        ("Frontend Development (React)", 120, 175.00, 0.0),
        ("Backend API Development (Python)", 80, 185.00, 0.0),
        ("Database Design & Migration", 1, 4500.00, 0.0),
        ("QA Testing & Bug Fixes", 60, 125.00, 0.0),
        ("DevOps Setup (CI/CD Pipeline)", 1, 3200.00, 0.0),
        ("Cloud Hosting Setup (AWS)", 1, 2800.00, 0.08),
        ("SSL Certificate (1 year)", 1, 199.00, 0.08),
        ("Domain Registration (2 years)", 1, 45.98, 0.08),
        ("Project Management", 160, 95.00, 0.0),
        ("Documentation & Training", 24, 125.00, 0.0),
        ("Post-Launch Support (3 months)", 1, 6000.00, 0.0),
        ("Emergency Hotfix Retainer", 1, 2500.00, 0.0),
        ("Performance Audit & Optimization", 1, 3800.00, 0.0),
    ]

    records: list[dict] = []
    formula_cache: dict[str, float | int | str] = {}
    for idx, (desc, qty, price, tax) in enumerate(line_items, 1):
        r = item_row + idx
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=qty)
        ws.cell(row=r, column=4, value=price)
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=tax)
        ws.cell(row=r, column=5).number_format = "0%"
        # ^ Line Total = Qty * Unit Price * (1 + Tax Rate)
        line_total = round(qty * price * (1 + tax), 2)
        ws.cell(row=r, column=6, value=f"=C{r}*D{r}*(1+E{r})")
        ws.cell(row=r, column=6).number_format = "#,##0.00"
        formula_cache[f"F{r}"] = line_total
        for col in range(1, 7):
            _style_data_cell(ws, r, col)
        records.append(
            {
                "line_number": idx,
                "description": desc,
                "quantity": qty,
                "unit_price": price,
                "tax_rate": tax,
                "line_total": line_total,
            }
        )

    last_item_row = item_row + len(line_items)
    summary_start = last_item_row + 2

    # ^ Pre-compute summary values
    subtotal = sum(qty * price for _, qty, price, _ in line_items)
    grand_total = sum(round(qty * price * (1 + tax), 2) for _, qty, price, tax in line_items)
    tax_total = round(grand_total - subtotal, 2)

    # * Subtotal
    ws.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=5)
    ws.cell(row=summary_start, column=1, value="Subtotal")
    ws.cell(row=summary_start, column=1).font = BOLD_FONT
    ws.cell(row=summary_start, column=1).alignment = Alignment(horizontal="right")
    ws.cell(
        row=summary_start,
        column=6,
        value=f"=SUMPRODUCT(C{item_row + 1}:C{last_item_row},D{item_row + 1}:D{last_item_row})",
    )
    ws.cell(row=summary_start, column=6).number_format = "#,##0.00"
    ws.cell(row=summary_start, column=6).font = BOLD_FONT
    formula_cache[f"F{summary_start}"] = subtotal

    # * Tax
    ws.merge_cells(
        start_row=summary_start + 1, start_column=1, end_row=summary_start + 1, end_column=5
    )
    ws.cell(row=summary_start + 1, column=1, value="Tax")
    ws.cell(row=summary_start + 1, column=1).alignment = Alignment(horizontal="right")
    ws.cell(
        row=summary_start + 1,
        column=6,
        value=f"=SUM(F{item_row + 1}:F{last_item_row})-F{summary_start}",
    )
    ws.cell(row=summary_start + 1, column=6).number_format = "#,##0.00"
    formula_cache[f"F{summary_start + 1}"] = tax_total

    # * Grand Total
    ws.merge_cells(
        start_row=summary_start + 2, start_column=1, end_row=summary_start + 2, end_column=5
    )
    ws.cell(row=summary_start + 2, column=1, value="GRAND TOTAL")
    ws.cell(row=summary_start + 2, column=1).font = Font(bold=True, size=13)
    ws.cell(row=summary_start + 2, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=summary_start + 2, column=6, value=f"=SUM(F{item_row + 1}:F{last_item_row})")
    ws.cell(row=summary_start + 2, column=6).number_format = "#,##0.00"
    ws.cell(row=summary_start + 2, column=6).font = Font(bold=True, size=13)
    formula_cache[f"F{summary_start + 2}"] = grand_total

    # * Notes section
    notes_row = summary_start + 5
    ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=6)
    ws["A" + str(notes_row)] = "Notes & Terms"
    ws["A" + str(notes_row)].font = SECTION_FONT

    notes = [
        "1. Payment is due within 30 days of invoice date.",
        "2. Late payments are subject to 1.5% monthly interest.",
        "3. All work is covered by a 90-day warranty period.",
        "4. Additional change requests will be billed at standard hourly rates.",
    ]
    for ni, note in enumerate(notes):
        r = notes_row + 1 + ni
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value=note)

    _auto_width(ws)
    ws.column_dimensions["B"].width = 40
    _save_with_formula_cache(
        wb, FIXTURES_DIR / "complex_invoice_form.xlsm", {"Invoice": formula_cache}
    )
    print("  Created complex_invoice_form.xlsm (form header + 15 line items + formulas)")
    return records


# * ──────────────────────────────────────────────────────────────────────
# * .xls files (legacy format via xlwt)
# * ──────────────────────────────────────────────────────────────────────


def create_legacy_inventory() -> list[dict]:
    """Basic .xls inventory with 20 rows."""
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Inventory")

    header_style = xlwt.easyxf(
        "font: bold on; align: horiz center; pattern: pattern solid, fore_colour light_blue;"
    )
    headers = [
        "SKU",
        "Product Name",
        "Category",
        "Qty In Stock",
        "Unit Cost",
        "Reorder Level",
        "Last Restocked",
    ]
    for col, h in enumerate(headers):
        ws.write(0, col, h, header_style)

    items = [
        ("SKU-1001", "Steel Bolt M8x50", "Fasteners", 5200, 0.12, 1000),
        ("SKU-1002", "Steel Nut M8", "Fasteners", 8400, 0.08, 2000),
        ("SKU-1003", "Copper Wire 14AWG (m)", "Electrical", 3200, 1.45, 500),
        ("SKU-1004", "PVC Pipe 2in (ft)", "Plumbing", 1800, 2.30, 300),
        ("SKU-1005", "Wood Screw #8x2", "Fasteners", 12000, 0.05, 3000),
        ("SKU-1006", "LED Bulb 60W", "Electrical", 450, 3.99, 100),
        ("SKU-1007", "Paint Brush 3in", "Tools", 320, 4.50, 50),
        ("SKU-1008", "Sandpaper 220 Grit", "Abrasives", 2800, 0.75, 500),
        ("SKU-1009", "Masking Tape 1in", "Adhesives", 1500, 2.10, 200),
        ("SKU-1010", "Drill Bit Set 29pc", "Tools", 85, 24.99, 20),
        ("SKU-1011", "Safety Goggles", "PPE", 340, 5.99, 50),
        ("SKU-1012", "Work Gloves (pair)", "PPE", 520, 8.50, 100),
        ("SKU-1013", "Measuring Tape 25ft", "Tools", 190, 12.99, 30),
        ("SKU-1014", "Caulk Silicone 10oz", "Adhesives", 680, 6.75, 100),
        ("SKU-1015", "Circuit Breaker 20A", "Electrical", 240, 8.25, 50),
        ("SKU-1016", "Pipe Fitting 2in Elbow", "Plumbing", 950, 3.15, 200),
        ("SKU-1017", "Concrete Mix 80lb", "Materials", 120, 5.50, 30),
        ("SKU-1018", "Rebar #4 (10ft)", "Materials", 350, 7.80, 50),
        ("SKU-1019", "Insulation R-13 Roll", "Materials", 95, 42.00, 20),
        ("SKU-1020", "Electrical Tape Black", "Electrical", 2200, 1.25, 500),
    ]

    records: list[dict] = []
    for i, (sku, name, cat, qty, cost, reorder) in enumerate(items):
        row = i + 1
        ws.write(row, 0, sku)
        ws.write(row, 1, name)
        ws.write(row, 2, cat)
        ws.write(row, 3, qty)
        ws.write(row, 4, cost)
        ws.write(row, 5, reorder)
        restock_date = _rand_date(date(2024, 1, 1), date(2024, 11, 30))
        ws.write(row, 6, restock_date.strftime("%Y-%m-%d"))
        records.append(
            {
                "sku": sku,
                "product_name": name,
                "category": cat,
                "qty_in_stock": qty,
                "unit_cost": cost,
                "reorder_level": reorder,
                "last_restocked": restock_date.isoformat(),
            }
        )

    for col in range(len(headers)):
        ws.col(col).width = 5000

    wb.save(str(FIXTURES_DIR / "legacy_inventory.xls"))
    print("  Created legacy_inventory.xls (20 rows)")
    return records


def create_legacy_payroll() -> list[dict]:
    """50-row .xls payroll dataset."""
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Payroll")

    header_style = xlwt.easyxf(
        "font: bold on; align: horiz center; pattern: pattern solid, fore_colour light_green;"
    )
    money_fmt = xlwt.easyxf(num_format_str="#,##0.00")

    headers = [
        "Employee ID",
        "Name",
        "Department",
        "Pay Period",
        "Base Salary",
        "Overtime Hours",
        "Overtime Pay",
        "Bonus",
        "Deductions",
        "Net Pay",
    ]
    for col, h in enumerate(headers):
        ws.write(0, col, h, header_style)

    records: list[dict] = []
    for i in range(1, 51):
        name = _rand_name()
        dept = random.choice(DEPARTMENTS)
        base = round(random.uniform(3500, 12000), 2)
        ot_hours = round(random.uniform(0, 20), 1)
        hourly = base / 160
        ot_pay = round(ot_hours * hourly * 1.5, 2)
        bonus = round(random.uniform(0, 2000), 2) if random.random() > 0.6 else 0
        deductions = round(base * random.uniform(0.15, 0.35), 2)
        net = round(base + ot_pay + bonus - deductions, 2)

        period_start = date(2024, 10, 1)
        period_end = date(2024, 10, 31)

        ws.write(i, 0, f"EMP-{2000 + i}")
        ws.write(i, 1, name)
        ws.write(i, 2, dept)
        ws.write(i, 3, f"{period_start.strftime('%Y-%m-%d')} to {period_end.strftime('%Y-%m-%d')}")
        ws.write(i, 4, base, money_fmt)
        ws.write(i, 5, ot_hours)
        ws.write(i, 6, ot_pay, money_fmt)
        ws.write(i, 7, bonus, money_fmt)
        ws.write(i, 8, deductions, money_fmt)
        ws.write(i, 9, net, money_fmt)

        records.append(
            {
                "employee_id": f"EMP-{2000 + i}",
                "name": name,
                "department": dept,
                "base_salary": base,
                "overtime_hours": ot_hours,
                "overtime_pay": ot_pay,
                "bonus": float(bonus),
                "deductions": deductions,
                "net_pay": net,
            }
        )

    for col in range(len(headers)):
        ws.col(col).width = 5500

    wb.save(str(FIXTURES_DIR / "legacy_payroll.xls"))
    print("  Created legacy_payroll.xls (50 rows)")
    return records


def create_legacy_student_grades() -> list[dict]:
    """40-row .xls student grade book."""
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Grades")

    header_style = xlwt.easyxf(
        "font: bold on; align: horiz center; pattern: pattern solid, fore_colour light_yellow;"
    )

    headers = [
        "Student ID",
        "Name",
        "Math",
        "English",
        "Science",
        "History",
        "Art",
        "Average",
        "Grade Letter",
        "Pass/Fail",
    ]
    for col, h in enumerate(headers):
        ws.write(0, col, h, header_style)

    def grade_letter(avg: float) -> str:
        if avg >= 90:
            return "A"
        if avg >= 80:
            return "B"
        if avg >= 70:
            return "C"
        if avg >= 60:
            return "D"
        return "F"

    records: list[dict] = []
    for i in range(1, 41):
        name = _rand_name()
        math = random.randint(35, 100)
        eng = random.randint(40, 100)
        sci = random.randint(30, 100)
        hist = random.randint(45, 100)
        art = random.randint(50, 100)
        avg = round((math + eng + sci + hist + art) / 5, 1)
        letter = grade_letter(avg)
        pf = "Pass" if avg >= 60 else "Fail"

        ws.write(i, 0, f"STU-{3000 + i}")
        ws.write(i, 1, name)
        ws.write(i, 2, math)
        ws.write(i, 3, eng)
        ws.write(i, 4, sci)
        ws.write(i, 5, hist)
        ws.write(i, 6, art)
        ws.write(i, 7, avg)
        ws.write(i, 8, letter)
        ws.write(i, 9, pf)

        records.append(
            {
                "student_id": f"STU-{3000 + i}",
                "name": name,
                "math": math,
                "english": eng,
                "science": sci,
                "history": hist,
                "art": art,
                "average": avg,
                "grade_letter": letter,
                "pass_fail": pf,
            }
        )

    for col in range(len(headers)):
        ws.col(col).width = 4500

    wb.save(str(FIXTURES_DIR / "legacy_student_grades.xls"))
    print("  Created legacy_student_grades.xls (40 rows)")
    return records


# * ──────────────────────────────────────────────────────────────────────
# * Edge-case / challenging structure files
# * ──────────────────────────────────────────────────────────────────────


def create_edge_sparse_pivot() -> list[dict]:
    """Cross-tab / pivot-style layout with sparse data and merged headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Pivot"

    ws.merge_cells("A1:N1")
    ws["A1"] = "Product Sales by Month (Units Sold)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    # * Column headers: months, merged into quarters
    quarters_cols = {
        "Q1 2024": ["Jan", "Feb", "Mar"],
        "Q2 2024": ["Apr", "May", "Jun"],
        "Q3 2024": ["Jul", "Aug", "Sep"],
        "Q4 2024": ["Oct", "Nov", "Dec"],
    }

    ws.cell(row=3, column=1, value="Category")
    ws.cell(row=4, column=1, value="Product")
    ws.merge_cells("A3:A4")
    ws.cell(row=3, column=1).font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=1).alignment = CENTER

    col = 2
    for q_name, month_list in quarters_cols.items():
        end_col = col + len(month_list) - 1
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=end_col)
        c = ws.cell(row=3, column=col, value=q_name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        for m in month_list:
            c2 = ws.cell(row=4, column=col, value=m)
            c2.font = Font(bold=True)
            c2.fill = SUBHEADER_FILL
            c2.alignment = CENTER
            col += 1

    # ^ "Total" column
    ws.cell(row=3, column=14, value="Total")
    ws.merge_cells(start_row=3, start_column=14, end_row=4, end_column=14)
    ws.cell(row=3, column=14).font = HEADER_FONT
    ws.cell(row=3, column=14).fill = HEADER_FILL
    ws.cell(row=3, column=14).alignment = CENTER

    # * Row data: categories with vertically merged labels
    product_groups = {
        "Hardware": [
            "Laptop Pro 15",
            "Desktop Tower",
            "Monitor 27in",
            "Server Rack Unit",
        ],
        "Software": [
            "Office Suite License",
            "Antivirus Annual",
            "Cloud Storage Plan",
        ],
        "Networking": [
            "Router Enterprise",
            "Switch 48-port",
            "Access Point",
            "Firewall Appliance",
            "Cat6 Cable Box",
        ],
        "Services": [
            "Installation Service",
            "Annual Maintenance",
            "Consulting (hourly)",
        ],
    }

    row = 5
    for cat, products in product_groups.items():
        for pi, prod in enumerate(products):
            ws.cell(row=row, column=1, value=prod if pi > 0 else prod)
            for mcol in range(2, 14):
                # ^ Sparse: some products not sold in some months (30% chance of empty)
                if random.random() > 0.3:
                    ws.cell(row=row, column=mcol, value=random.randint(1, 500))
                # ^ else: leave cell empty (sparse data)
            # ^ Total formula
            cl_start = get_column_letter(2)
            cl_end = get_column_letter(13)
            ws.cell(row=row, column=14, value=f"=SUM({cl_start}{row}:{cl_end}{row})")
            ws.cell(row=row, column=14).font = BOLD_FONT
            row += 1

        # ^ Merge category cells vertically
        if len(products) > 1:
            # ^ Insert category label in a separate column (col A becomes product, add category in structure)
            pass
        # ^ Actually, let's add a category column approach instead
        row += 1  # ^ blank row between groups

    # ^ Redo: use column A as category (merged) and column B as product
    # ^ Let me rebuild with proper structure
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Sales Pivot"

    ws2.merge_cells("A1:O1")
    ws2["A1"] = "Product Sales by Month (Units Sold)"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].alignment = CENTER

    # * Headers
    ws2.cell(row=3, column=1, value="Category")
    ws2.cell(row=3, column=2, value="Product")
    ws2.merge_cells("A3:A4")
    ws2.merge_cells("B3:B4")
    for c in [1, 2]:
        ws2.cell(row=3, column=c).font = HEADER_FONT
        ws2.cell(row=3, column=c).fill = HEADER_FILL
        ws2.cell(row=3, column=c).alignment = CENTER

    col = 3
    for q_name, month_list in quarters_cols.items():
        end_col = col + len(month_list) - 1
        ws2.merge_cells(start_row=3, start_column=col, end_row=3, end_column=end_col)
        c = ws2.cell(row=3, column=col, value=q_name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        for m in month_list:
            c2 = ws2.cell(row=4, column=col, value=m)
            c2.font = Font(bold=True)
            c2.fill = SUBHEADER_FILL
            c2.alignment = CENTER
            col += 1

    ws2.cell(row=3, column=15, value="Annual Total")
    ws2.merge_cells(start_row=3, start_column=15, end_row=4, end_column=15)
    ws2.cell(row=3, column=15).font = HEADER_FONT
    ws2.cell(row=3, column=15).fill = HEADER_FILL
    ws2.cell(row=3, column=15).alignment = CENTER

    # ^ Month name list in order (columns 3..14)
    month_names_ordered = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]

    records: list[dict] = []
    formula_cache: dict[str, float | int | str] = {}
    row = 5
    for cat, products in product_groups.items():
        cat_start_row = row
        for pi, prod in enumerate(products):
            ws2.cell(row=row, column=2, value=prod)
            row_units_total = 0
            for mcol in range(3, 15):
                if random.random() > 0.25:
                    units = random.randint(1, 500)
                    ws2.cell(row=row, column=mcol, value=units)
                    month_name = month_names_ordered[mcol - 3]
                    row_units_total += units
                    records.append(
                        {
                            "category": cat,
                            "product": prod,
                            "month": month_name,
                            "units_sold": units,
                        }
                    )
            c_s = get_column_letter(3)
            c_e = get_column_letter(14)
            ws2.cell(row=row, column=15, value=f"=SUM({c_s}{row}:{c_e}{row})")
            ws2.cell(row=row, column=15).font = BOLD_FONT
            formula_cache[f"O{row}"] = row_units_total
            row += 1

        cat_end_row = row - 1
        if cat_end_row > cat_start_row:
            ws2.merge_cells(
                start_row=cat_start_row,
                start_column=1,
                end_row=cat_end_row,
                end_column=1,
            )
        ws2.cell(row=cat_start_row, column=1, value=cat)
        ws2.cell(row=cat_start_row, column=1).font = BOLD_FONT
        ws2.cell(row=cat_start_row, column=1).alignment = Alignment(
            vertical="center",
            horizontal="center",
            text_rotation=0,
        )

    _auto_width(ws2)
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 25
    _save_with_formula_cache(
        wb2, FIXTURES_DIR / "edge_sparse_pivot.xlsx", {"Sales Pivot": formula_cache}
    )
    print("  Created edge_sparse_pivot.xlsx (pivot-style, sparse data, vertical merges)")
    return records


def create_edge_mixed_layout() -> list[dict]:
    """Mixed layout: form header → empty row → table data → notes section."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Report"

    # * Company banner (merged)
    ws.merge_cells("A1:G2")
    ws["A1"] = "GLOBALTECH SOLUTIONS"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = CENTER

    # * Report metadata (key-value pairs)
    kv_pairs = [
        ("Report Type:", "Monthly Order Summary"),
        ("Period:", "October 2024"),
        ("Generated By:", "System Admin"),
        ("Report ID:", "RPT-2024-10-0042"),
        ("Confidentiality:", "Internal Use Only"),
    ]
    for ki, (k, v) in enumerate(kv_pairs):
        row = 4 + ki
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws.cell(row=row, column=3, value=v)

    # * Empty separator row 9
    # * Section header
    ws.merge_cells("A10:G10")
    ws["A10"] = "ORDER DETAILS"
    ws["A10"].font = SECTION_FONT
    ws["A10"].fill = HEADER_FILL
    ws["A10"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A10"].alignment = CENTER

    # * Table headers (row 11)
    t_headers = ["Order ID", "Customer", "Product", "Qty", "Unit Price", "Total", "Status"]
    for col, h in enumerate(t_headers, 1):
        ws.cell(row=11, column=col, value=h)
    _style_header_row(ws, 11, 1, len(t_headers))

    # * 25 order rows
    statuses = [
        "Shipped",
        "Shipped",
        "Shipped",
        "Processing",
        "Delivered",
        "Delivered",
        "Cancelled",
    ]
    records: list[dict] = []
    formula_cache: dict[str, float | int | str] = {}
    grand_total_sum = 0.0
    for i in range(25):
        row = 12 + i
        qty = random.randint(1, 50)
        price = round(random.uniform(19.99, 999.99), 2)
        customer = _rand_name()
        product = random.choice(PRODUCTS)
        status = random.choice(statuses)
        line_total = round(qty * price, 2)

        ws.cell(row=row, column=1, value=f"ORD-{50000 + i}")
        ws.cell(row=row, column=2, value=customer)
        ws.cell(row=row, column=3, value=product)
        ws.cell(row=row, column=4, value=qty)
        ws.cell(row=row, column=5, value=price)
        ws.cell(row=row, column=5).number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        ws.cell(row=row, column=6).number_format = "#,##0.00"
        formula_cache[f"F{row}"] = line_total
        ws.cell(row=row, column=7, value=status)
        for col in range(1, 8):
            _style_data_cell(ws, row, col)

        grand_total_sum += line_total
        records.append(
            {
                "order_id": f"ORD-{50000 + i}",
                "customer": customer,
                "product": product,
                "quantity": qty,
                "unit_price": price,
                "total": line_total,
                "status": status,
            }
        )

    last_data_row = 12 + 24

    # * Summary row
    summary_row = last_data_row + 1
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=1).font = BOLD_FONT
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=summary_row, column=6, value=f"=SUM(F12:F{last_data_row})")
    ws.cell(row=summary_row, column=6).number_format = "#,##0.00"
    ws.cell(row=summary_row, column=6).font = BOLD_FONT
    formula_cache[f"F{summary_row}"] = round(grand_total_sum, 2)

    # * Empty separator
    # * Notes section
    notes_row = summary_row + 3
    ws.merge_cells(f"A{notes_row}:G{notes_row}")
    ws.cell(row=notes_row, column=1, value="NOTES & COMMENTS")
    ws.cell(row=notes_row, column=1).font = SECTION_FONT

    notes = [
        "All orders above $500 require manager approval before processing.",
        "Cancelled orders from Q4 are eligible for reactivation until Dec 31.",
        "Bulk discount of 15% applies to orders with quantity >= 20 units.",
        "Next review meeting scheduled for Nov 5, 2024 at 2:00 PM EST.",
    ]
    for ni, note in enumerate(notes):
        r = notes_row + 1 + ni
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(row=r, column=1, value=note)
        ws.cell(row=r, column=1).alignment = WRAP

    _auto_width(ws)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    _save_with_formula_cache(
        wb, FIXTURES_DIR / "edge_mixed_layout.xlsx", {"Order Report": formula_cache}
    )
    print("  Created edge_mixed_layout.xlsx (form header + table + notes)")
    return records


def create_edge_nested_groups() -> list[dict]:
    """3-level hierarchical grouping with vertical merges."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organization"

    ws.merge_cells("A1:H1")
    ws["A1"] = "Company Organization Directory"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    headers = [
        "Department",
        "Team",
        "Sub-Team",
        "Employee ID",
        "Name",
        "Title",
        "Location",
        "Start Date",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header_row(ws, 3, 1, len(headers))

    org = {
        "Engineering": {
            "Platform": {
                "Backend": [
                    ("Sr. Software Engineer", "San Francisco"),
                    ("Software Engineer", "San Francisco"),
                    ("Software Engineer", "Austin"),
                    ("Jr. Software Engineer", "Remote"),
                ],
                "Infrastructure": [
                    ("Staff Engineer", "San Francisco"),
                    ("DevOps Engineer", "Seattle"),
                    ("SRE", "Seattle"),
                ],
            },
            "Product": {
                "Frontend": [
                    ("Sr. Frontend Engineer", "New York"),
                    ("Frontend Engineer", "New York"),
                    ("Frontend Engineer", "Remote"),
                ],
                "Mobile": [
                    ("iOS Engineer", "San Francisco"),
                    ("Android Engineer", "San Francisco"),
                    ("QA Engineer", "Austin"),
                ],
            },
            "Data": {
                "Analytics": [
                    ("Data Analyst", "New York"),
                    ("Sr. Data Analyst", "New York"),
                ],
                "ML Engineering": [
                    ("ML Engineer", "San Francisco"),
                    ("Sr. ML Engineer", "San Francisco"),
                    ("Research Scientist", "Remote"),
                ],
            },
        },
        "Sales": {
            "Enterprise": {
                "North America": [
                    ("VP Sales", "New York"),
                    ("Account Executive", "New York"),
                    ("Account Executive", "Chicago"),
                    ("Sales Engineer", "Austin"),
                ],
                "EMEA": [
                    ("Regional Director", "London"),
                    ("Account Executive", "London"),
                    ("Account Executive", "Berlin"),
                ],
            },
            "SMB": {
                "Inside Sales": [
                    ("Sales Rep", "Austin"),
                    ("Sales Rep", "Austin"),
                    ("Sales Rep", "Remote"),
                    ("SDR", "Remote"),
                    ("SDR", "Remote"),
                ],
                "Partnerships": [
                    ("Partner Manager", "San Francisco"),
                    ("Partner Manager", "New York"),
                ],
            },
        },
        "Marketing": {
            "Growth": {
                "Digital": [
                    ("Growth Marketing Manager", "San Francisco"),
                    ("SEO Specialist", "Remote"),
                    ("Paid Ads Specialist", "New York"),
                ],
                "Content": [
                    ("Content Lead", "New York"),
                    ("Technical Writer", "Remote"),
                    ("Copywriter", "New York"),
                    ("Video Producer", "Los Angeles"),
                ],
            },
            "Brand": {
                "Design": [
                    ("Creative Director", "New York"),
                    ("Sr. Designer", "New York"),
                    ("Designer", "Remote"),
                ],
                "Events": [
                    ("Events Manager", "San Francisco"),
                    ("Events Coordinator", "Austin"),
                ],
            },
        },
    }

    emp_id = 5001
    row = 4
    dept_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    team_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    records: list[dict] = []
    for dept_name, teams in org.items():
        dept_start = row
        for team_name, sub_teams in teams.items():
            team_start = row
            for sub_name, members in sub_teams.items():
                sub_start = row
                for title, location in members:
                    name = _rand_name()
                    hire = _rand_date(date(2019, 1, 1), date(2024, 10, 31))
                    ws.cell(row=row, column=4, value=f"EMP-{emp_id}")
                    ws.cell(row=row, column=5, value=name)
                    ws.cell(row=row, column=6, value=title)
                    ws.cell(row=row, column=7, value=location)
                    ws.cell(row=row, column=8, value=hire)
                    ws.cell(row=row, column=8).number_format = "YYYY-MM-DD"
                    records.append(
                        {
                            "department": dept_name,
                            "team": team_name,
                            "sub_team": sub_name,
                            "employee_id": f"EMP-{emp_id}",
                            "name": name,
                            "title": title,
                            "location": location,
                            "start_date": hire.isoformat(),
                        }
                    )
                    emp_id += 1
                    row += 1

                sub_end = row - 1
                if sub_end > sub_start:
                    ws.merge_cells(
                        start_row=sub_start,
                        start_column=3,
                        end_row=sub_end,
                        end_column=3,
                    )
                ws.cell(row=sub_start, column=3, value=sub_name)
                ws.cell(row=sub_start, column=3).alignment = Alignment(vertical="center")

            team_end = row - 1
            if team_end > team_start:
                ws.merge_cells(
                    start_row=team_start,
                    start_column=2,
                    end_row=team_end,
                    end_column=2,
                )
            ws.cell(row=team_start, column=2, value=team_name)
            ws.cell(row=team_start, column=2).font = BOLD_FONT
            ws.cell(row=team_start, column=2).alignment = Alignment(vertical="center")
            ws.cell(row=team_start, column=2).fill = team_fill

        dept_end = row - 1
        if dept_end > dept_start:
            ws.merge_cells(
                start_row=dept_start,
                start_column=1,
                end_row=dept_end,
                end_column=1,
            )
        ws.cell(row=dept_start, column=1, value=dept_name)
        ws.cell(row=dept_start, column=1).font = Font(bold=True, size=11)
        ws.cell(row=dept_start, column=1).alignment = Alignment(vertical="center")
        ws.cell(row=dept_start, column=1).fill = dept_fill

    # * Headcount summary at the bottom
    row += 2
    ws.cell(row=row, column=1, value="Total Headcount:")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=2, value=emp_id - 5001)
    ws.cell(row=row, column=2).font = BOLD_FONT

    _auto_width(ws)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["F"].width = 28
    wb.save(FIXTURES_DIR / "edge_nested_groups.xlsx")
    print(f"  Created edge_nested_groups.xlsx (3-level vertical merges, {emp_id - 5001} employees)")
    return records


# * ──────────────────────────────────────────────────────────────────────
# * Main
# * ──────────────────────────────────────────────────────────────────────


def main():
    print(f"Generating test fixtures in: {FIXTURES_DIR}\n")

    print("[.xlsx - Simple / Merged / Large]")
    emp_records = create_simple_employee_directory()
    fin_records = create_merged_financial_report()
    txn_records = create_large_transaction_log()

    print("\n[.xlsm - Formulas / Multi-level Headers / Invoice Form]")
    budget_records = create_formula_budget_tracker()
    sales_records = create_multi_level_header_sales()
    invoice_records = create_complex_invoice_form()

    print("\n[.xls - Legacy Format]")
    inv_records = create_legacy_inventory()
    payroll_records = create_legacy_payroll()
    grade_records = create_legacy_student_grades()

    print("\n[Edge Cases - Sparse Pivot / Mixed Layout / Nested Groups]")
    pivot_records = create_edge_sparse_pivot()
    order_records = create_edge_mixed_layout()
    org_records = create_edge_nested_groups()

    print(f"\nSaving expected JSON to: {EXPECTED_DIR}")
    _save_expected("simple_employee_directory.json", emp_records)
    _save_expected("merged_financial_report.json", fin_records)
    _save_expected("large_transaction_log.json", txn_records)
    _save_expected("formula_budget_tracker.json", budget_records)
    _save_expected("multi_level_header_sales.json", sales_records)
    _save_expected("complex_invoice_form.json", invoice_records)
    _save_expected("legacy_inventory.json", inv_records)
    _save_expected("legacy_payroll.json", payroll_records)
    _save_expected("legacy_student_grades.json", grade_records)
    _save_expected("edge_sparse_pivot.json", pivot_records)
    _save_expected("edge_mixed_layout.json", order_records)
    _save_expected("edge_nested_groups.json", org_records)

    data_files = len(list(FIXTURES_DIR.iterdir()))
    expected_files = len(list(EXPECTED_DIR.iterdir()))
    print(f"\nDone! {data_files} Excel files, {expected_files} expected JSON files generated.")


if __name__ == "__main__":
    main()
