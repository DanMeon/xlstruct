"""Tests for HybridReader (the sole Excel reader)."""

import io
from datetime import date, datetime

import openpyxl
import pytest
import xlwt

from xlstruct.exceptions import ReaderError
from xlstruct.reader.hybrid_reader import HybridReader

# * .xls fixtures


@pytest.fixture
def simple_xls_bytes() -> bytes:
    """Generate a real .xls file in memory."""
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Products")
    ws.write(0, 0, "Name")
    ws.write(0, 1, "Price")
    ws.write(0, 2, "Stock")
    ws.write(1, 0, "Apple")
    ws.write(1, 1, 1.5)
    ws.write(1, 2, 100)
    ws.write(2, 0, "Banana")
    ws.write(2, 1, 0.75)
    ws.write(2, 2, 200)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def merged_xls_bytes() -> bytes:
    """Generate a .xls file with merged cells."""
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Invoice")
    ws.write_merge(0, 0, 0, 3, "Invoice #2024-001")
    ws.write(1, 0, "Item")
    ws.write(1, 1, "Qty")
    ws.write(1, 2, "Price")
    ws.write(1, 3, "Total")
    ws.write(2, 0, "Widget")
    ws.write(2, 1, 10)
    ws.write(2, 2, 25.0)
    ws.write(2, 3, 250.0)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# * .xlsx fixtures for merged cells


@pytest.fixture
def merged_xlsx_bytes() -> bytes:
    """Generate a .xlsx file with merged cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Invoice #2024-001"
    ws["A2"] = "Item"
    ws["B2"] = "Qty"
    ws["C2"] = "Price"
    ws["D2"] = "Total"
    ws["A3"] = "Widget"
    ws["B3"] = 10
    ws["C3"] = 25.0
    ws["D3"] = 250.0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# * HybridReader — .xlsx tests


class TestHybridReader:
    """HybridReader must produce correct output from xlsx files."""

    def test_read_simple(self, simple_xlsx_bytes: bytes):
        reader = HybridReader()
        wb = reader.read(simple_xlsx_bytes)

        assert len(wb.sheets) == 1
        assert wb.sheets[0].name == "Products"

        sheet = wb.sheets[0]
        assert sheet.row_count > 0
        assert sheet.col_count > 0

        header_a = sheet.get_cell(1, 1)
        assert header_a is not None
        assert header_a.display_value is not None

    def test_read_specific_sheet(self, simple_xlsx_bytes: bytes):
        reader = HybridReader()
        wb = reader.read(simple_xlsx_bytes, sheet_name="Products")
        assert len(wb.sheets) == 1
        assert wb.sheets[0].name == "Products"

    def test_read_nonexistent_sheet(self, simple_xlsx_bytes: bytes):
        reader = HybridReader()
        with pytest.raises(ReaderError, match="not found"):
            reader.read(simple_xlsx_bytes, sheet_name="DoesNotExist")

    def test_read_invalid_bytes(self):
        reader = HybridReader()
        with pytest.raises(ReaderError):
            reader.read(b"not a valid xlsx file")

    def test_merged_cells(self, merged_xlsx_bytes: bytes):
        reader = HybridReader()
        wb = reader.read(merged_xlsx_bytes)
        sheet = wb.sheets[0]

        assert len(sheet.merged_ranges) == 1
        assert "A1:D1" in sheet.merged_ranges

        # ^ Origin cell
        origin = sheet.get_cell(1, 1)
        assert origin is not None
        assert origin.value == "Invoice #2024-001"
        assert origin.is_merged is True
        assert origin.merge_origin is None

        # ^ Non-origin merged cell — not included (empty cells skipped)
        non_origin = sheet.get_cell(1, 2)
        assert non_origin is None

    def test_date_cells(self):
        """calamine returns datetime.date for date-only cells."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dates"
        ws["A1"] = "Date"
        ws["A2"] = date(2024, 1, 15)
        ws["A3"] = datetime(2024, 6, 30, 14, 30, 0)
        buf = io.BytesIO()
        wb.save(buf)

        reader = HybridReader()
        result = reader.read(buf.getvalue())
        sheet = result.sheets[0]

        date_cell = sheet.get_cell(2, 1)
        assert date_cell is not None
        assert date_cell.data_type == "d"

    def test_boolean_cells(self):
        """calamine returns native bool."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bools"
        ws["A1"] = "Flag"
        ws["A2"] = True
        ws["A3"] = False
        buf = io.BytesIO()
        wb.save(buf)

        reader = HybridReader()
        result = reader.read(buf.getvalue())
        sheet = result.sheets[0]

        true_cell = sheet.get_cell(2, 1)
        assert true_cell is not None
        assert true_cell.value is True
        assert true_cell.data_type == "b"

        false_cell = sheet.get_cell(3, 1)
        assert false_cell is not None
        assert false_cell.value is False
        assert false_cell.data_type == "b"

    def test_empty_cells_skipped(self):
        """Empty cells should not appear in output."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sparse"
        ws["A1"] = "Name"
        ws["C1"] = "Value"  # ^ B1 is empty
        ws["A2"] = "X"
        ws["C2"] = 42
        buf = io.BytesIO()
        wb.save(buf)

        reader = HybridReader()
        result = reader.read(buf.getvalue())
        sheet = result.sheets[0]

        assert sheet.get_cell(1, 2) is None
        assert sheet.get_cell(2, 2) is None


# * HybridReader — .xls tests (Pass 2 skipped)


class TestHybridReaderXls:
    """HybridReader .xls support — calamine only, no formula extraction."""

    def test_read_simple(self, simple_xls_bytes: bytes):
        reader = HybridReader()
        wb = reader.read(simple_xls_bytes, source_ext=".xls")

        assert len(wb.sheets) == 1
        assert wb.sheets[0].name == "Products"

        sheet = wb.sheets[0]
        assert sheet.row_count == 3
        assert sheet.col_count == 3

        header_a = sheet.get_cell(1, 1)
        assert header_a is not None
        assert header_a.value == "Name"
        assert header_a.data_type == "s"

        price = sheet.get_cell(2, 2)
        assert price is not None
        assert price.value == 1.5
        assert price.data_type == "n"

    def test_read_specific_sheet(self, simple_xls_bytes: bytes):
        reader = HybridReader()
        wb = reader.read(simple_xls_bytes, sheet_name="Products", source_ext=".xls")
        assert len(wb.sheets) == 1
        assert wb.sheets[0].name == "Products"

    def test_read_nonexistent_sheet(self, simple_xls_bytes: bytes):
        reader = HybridReader()
        with pytest.raises(ReaderError, match="not found"):
            reader.read(simple_xls_bytes, sheet_name="DoesNotExist", source_ext=".xls")

    def test_merged_cells(self, merged_xls_bytes: bytes):
        reader = HybridReader()
        wb = reader.read(merged_xls_bytes, source_ext=".xls")
        sheet = wb.sheets[0]

        assert len(sheet.merged_ranges) == 1

        origin = sheet.get_cell(1, 1)
        assert origin is not None
        assert origin.value == "Invoice #2024-001"
        assert origin.is_merged is True

    def test_no_formula_extraction(self, simple_xls_bytes: bytes):
        """.xls: Pass 2 skipped, so formula is always None."""
        reader = HybridReader()
        wb = reader.read(simple_xls_bytes, source_ext=".xls")
        for cell in wb.sheets[0].cells:
            assert cell.formula is None

    def test_read_invalid_bytes(self):
        reader = HybridReader()
        with pytest.raises(ReaderError):
            reader.read(b"not a valid xls file", source_ext=".xls")


# * Extractor._get_source_ext() tests


class TestGetSourceExt:
    """Test Extractor._get_source_ext()."""

    def test_xlsx(self):
        from xlstruct.extractor import Extractor

        assert Extractor._get_source_ext("report.xlsx") == ".xlsx"

    def test_xls(self):
        from xlstruct.extractor import Extractor

        assert Extractor._get_source_ext("report.xls") == ".xls"

    def test_xlsm(self):
        from xlstruct.extractor import Extractor

        assert Extractor._get_source_ext("macro_book.xlsm") == ".xlsm"

    def test_unsupported(self):
        from xlstruct.extractor import Extractor

        with pytest.raises(ReaderError):
            Extractor._get_source_ext("data.pdf")

    def test_xls_in_url(self):
        from xlstruct.extractor import Extractor

        assert Extractor._get_source_ext("s3://bucket/data.xls") == ".xls"

    def test_xlsx_with_query_params(self):
        from xlstruct.extractor import Extractor

        assert Extractor._get_source_ext("https://example.com/file.xlsx?token=abc") == ".xlsx"
