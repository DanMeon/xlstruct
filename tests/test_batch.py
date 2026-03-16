"""Tests for batch extraction (extract_batch / extract_batch_sync)."""

import io
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest
from pydantic import BaseModel

from xlstruct.extractor import Extractor
from xlstruct.schemas.batch import BatchResult, FileResult
from xlstruct.schemas.usage import TokenUsage

# * Test schemas

class Product(BaseModel):
    name: str
    price: float


# * Fixtures

@pytest.fixture(autouse=True)
def _mock_instructor():
    """Mock ExtractionEngine._build_client to avoid real API calls."""
    with patch(
        "xlstruct.extraction.engine.ExtractionEngine._build_client",
        return_value=MagicMock(),
    ):
        yield


def _make_xlsx(tmp_path, filename: str, rows: list[tuple]) -> str:
    """Create a simple xlsx file and return its path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Price"
    for i, (name, price) in enumerate(rows, start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = price
    buf = io.BytesIO()
    wb.save(buf)
    path = tmp_path / filename
    path.write_bytes(buf.getvalue())
    return str(path)


@pytest.fixture
def xlsx_files(tmp_path) -> list[str]:
    """Three xlsx files with product data."""
    return [
        _make_xlsx(tmp_path, "products_1.xlsx", [("Apple", 1.5), ("Banana", 0.75)]),
        _make_xlsx(tmp_path, "products_2.xlsx", [("Cherry", 3.0)]),
        _make_xlsx(tmp_path, "products_3.xlsx", [("Durian", 10.0), ("Elderberry", 5.0)]),
    ]


# * Model tests

class TestBatchResultModel:
    def test_empty_batch(self):
        result = BatchResult(results=[])
        assert result.succeeded == 0
        assert result.failed == 0
        assert result.total == 0
        assert len(result) == 0
        assert result.all_records == []

    def test_mixed_results(self):
        results = BatchResult[Product](results=[
            FileResult(
                source="a.xlsx",
                success=True,
                records=[Product(name="Apple", price=1.5)],
                usage=TokenUsage(llm_calls=1, input_tokens=100, output_tokens=50, total_tokens=150),
            ),
            FileResult(
                source="b.xlsx",
                success=False,
                error="FileNotFoundError: b.xlsx",
            ),
            FileResult(
                source="c.xlsx",
                success=True,
                records=[Product(name="Cherry", price=3.0), Product(name="Date", price=4.0)],
                usage=TokenUsage(llm_calls=1, input_tokens=120, output_tokens=60, total_tokens=180),
            ),
        ])

        assert results.succeeded == 2
        assert results.failed == 1
        assert results.total == 3
        assert len(results) == 3

    def test_total_usage_aggregation(self):
        results = BatchResult[Product](results=[
            FileResult(
                source="a.xlsx",
                success=True,
                records=[],
                usage=TokenUsage(
                    llm_calls=1, input_tokens=100, output_tokens=50,
                    total_tokens=150, cache_creation_tokens=10, cache_read_tokens=5,
                ),
            ),
            FileResult(
                source="b.xlsx",
                success=True,
                records=[],
                usage=TokenUsage(
                    llm_calls=2, input_tokens=200, output_tokens=80,
                    total_tokens=280, cache_creation_tokens=0, cache_read_tokens=20,
                ),
            ),
        ])

        usage = results.total_usage
        assert usage.llm_calls == 3
        assert usage.input_tokens == 300
        assert usage.output_tokens == 130
        assert usage.total_tokens == 430
        assert usage.cache_creation_tokens == 10
        assert usage.cache_read_tokens == 25

    def test_total_usage_skips_none(self):
        """Failed files with no usage should not break aggregation."""
        results = BatchResult[Product](results=[
            FileResult(source="a.xlsx", success=False, error="err"),
            FileResult(
                source="b.xlsx",
                success=True,
                records=[],
                usage=TokenUsage(llm_calls=1, input_tokens=50, output_tokens=25, total_tokens=75),
            ),
        ])
        assert results.total_usage.llm_calls == 1

    def test_all_records(self):
        results = BatchResult[Product](results=[
            FileResult(
                source="a.xlsx",
                success=True,
                records=[Product(name="Apple", price=1.5)],
            ),
            FileResult(source="b.xlsx", success=False, error="err"),
            FileResult(
                source="c.xlsx",
                success=True,
                records=[Product(name="Cherry", price=3.0), Product(name="Date", price=4.0)],
            ),
        ])

        all_records = results.all_records
        assert len(all_records) == 3
        assert all_records[0].name == "Apple"
        assert all_records[2].name == "Date"

    def test_iteration(self):
        results = BatchResult[Product](results=[
            FileResult(source="a.xlsx", success=True, records=[]),
            FileResult(source="b.xlsx", success=True, records=[]),
        ])
        sources = [r.source for r in results]
        assert sources == ["a.xlsx", "b.xlsx"]

    def test_indexing(self):
        results = BatchResult[Product](results=[
            FileResult(source="a.xlsx", success=True, records=[]),
            FileResult(source="b.xlsx", success=True, records=[]),
        ])
        assert results[0].source == "a.xlsx"
        assert results[1].source == "b.xlsx"


# * Extractor integration tests

class TestExtractBatch:
    async def test_batch_all_succeed(self, xlsx_files):
        """All files succeed — records and usage returned."""
        extractor = Extractor()

        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.side_effect = [
                [Product(name="Apple", price=1.5), Product(name="Banana", price=0.75)],
                [Product(name="Cherry", price=3.0)],
                [Product(name="Durian", price=10.0), Product(name="Elderberry", price=5.0)],
            ]
            result = await extractor.extract_batch(xlsx_files, Product)

        assert result.succeeded == 3
        assert result.failed == 0
        assert len(result.all_records) == 5
        assert mock_extract.call_count == 3

    async def test_batch_partial_failure(self, xlsx_files):
        """One file fails — other files still processed."""
        extractor = Extractor()

        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.side_effect = [
                [Product(name="Apple", price=1.5)],
                ValueError("LLM error"),
                [Product(name="Durian", price=10.0)],
            ]
            result = await extractor.extract_batch(xlsx_files, Product)

        assert result.succeeded == 2
        assert result.failed == 1
        assert len(result.all_records) == 2

        # ^ Failed file has error message
        failed = [r for r in result if not r.success]
        assert len(failed) == 1
        assert "ValueError" in failed[0].error

    async def test_batch_all_fail(self, xlsx_files):
        """All files fail — empty records, all errors captured."""
        extractor = Extractor()

        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.side_effect = RuntimeError("API down")
            result = await extractor.extract_batch(xlsx_files, Product)

        assert result.succeeded == 0
        assert result.failed == 3
        assert result.all_records == []

    async def test_batch_empty_sources(self):
        """Empty source list returns empty result."""
        extractor = Extractor()
        result = await extractor.extract_batch([], Product)

        assert result.total == 0
        assert result.succeeded == 0

    async def test_batch_concurrency_respected(self, xlsx_files):
        """Concurrency limit is enforced via semaphore."""
        import asyncio

        max_concurrent = 0
        current_concurrent = 0
        lock = asyncio.Lock()

        original_extract = AsyncMock(return_value=[Product(name="X", price=1.0)])

        async def tracking_extract(*args, **kwargs):
            nonlocal max_concurrent, current_concurrent
            async with lock:
                current_concurrent += 1
                max_concurrent = max(max_concurrent, current_concurrent)
            try:
                return await original_extract(*args, **kwargs)
            finally:
                async with lock:
                    current_concurrent -= 1

        extractor = Extractor()

        with patch.object(extractor._engine, "extract", side_effect=tracking_extract):
            await extractor.extract_batch(xlsx_files, Product, concurrency=2)

        assert max_concurrent <= 2

    async def test_batch_preserves_source_order(self, xlsx_files):
        """Results are returned in the same order as sources."""
        extractor = Extractor()

        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = []
            result = await extractor.extract_batch(xlsx_files, Product)

        result_sources = [r.source for r in result]
        assert result_sources == xlsx_files


class TestExtractBatchSync:
    def test_sync_wrapper(self, xlsx_files):
        """Sync wrapper works correctly."""
        extractor = Extractor()

        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = [Product(name="X", price=1.0)]
            result = extractor.extract_batch_sync(xlsx_files, Product)

        assert result.succeeded == 3
        assert result.total == 3
