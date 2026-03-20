"""Tests for multi-sheet extraction (extract_workbook / extract_workbook_sync)."""

import io
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest
from pydantic import BaseModel

from xlstruct.extractor import Extractor
from xlstruct.schemas.usage import TokenUsage
from xlstruct.schemas.workbook import SheetResult, WorkbookResult

# * Test schemas

class SalesRow(BaseModel):
    product: str
    revenue: float


class ExpenseRow(BaseModel):
    category: str
    amount: float


class InventoryRow(BaseModel):
    item: str
    quantity: int


# * Fixtures

@pytest.fixture(autouse=True)
def _mock_instructor():
    """Mock build_instructor_client to avoid real API calls."""
    with patch(
        "xlstruct.extraction.engine.build_instructor_client",
        return_value=MagicMock(),
    ):
        yield


def _make_multi_sheet_xlsx(tmp_path) -> str:
    """Create an xlsx file with 3 sheets."""
    wb = openpyxl.Workbook()

    # * Sales sheet
    ws_sales = wb.active
    ws_sales.title = "Sales"
    ws_sales["A1"] = "Product"
    ws_sales["B1"] = "Revenue"
    ws_sales["A2"] = "Widget"
    ws_sales["B2"] = 1000.0

    # * Expenses sheet
    ws_expenses = wb.create_sheet("Expenses")
    ws_expenses["A1"] = "Category"
    ws_expenses["B1"] = "Amount"
    ws_expenses["A2"] = "Office"
    ws_expenses["B2"] = 500.0

    # * Inventory sheet
    ws_inventory = wb.create_sheet("Inventory")
    ws_inventory["A1"] = "Item"
    ws_inventory["B1"] = "Quantity"
    ws_inventory["A2"] = "Widget"
    ws_inventory["B2"] = 100

    buf = io.BytesIO()
    wb.save(buf)
    path = tmp_path / "report.xlsx"
    path.write_bytes(buf.getvalue())
    return str(path)


@pytest.fixture
def multi_sheet_xlsx(tmp_path) -> str:
    return _make_multi_sheet_xlsx(tmp_path)


# * Model tests

class TestWorkbookResultModel:
    def test_empty_result(self):
        result = WorkbookResult(results={})
        assert result.succeeded == 0
        assert result.failed == 0
        assert result.total == 0
        assert len(result) == 0
        assert result.sheet_names == []

    def test_mixed_results(self):
        result = WorkbookResult(results={
            "Sales": SheetResult(
                sheet_name="Sales",
                success=True,
                records=[SalesRow(product="Widget", revenue=1000.0)],
                usage=TokenUsage(
                    llm_calls=1, input_tokens=100, output_tokens=50, total_tokens=150
                ),
            ),
            "Expenses": SheetResult(
                sheet_name="Expenses",
                success=False,
                error="ValueError: parse error",
            ),
        })

        assert result.succeeded == 1
        assert result.failed == 1
        assert result.total == 2

    def test_total_usage_aggregation(self):
        result = WorkbookResult(results={
            "A": SheetResult(
                sheet_name="A",
                success=True,
                records=[],
                usage=TokenUsage(llm_calls=1, input_tokens=100, output_tokens=50, total_tokens=150),
            ),
            "B": SheetResult(
                sheet_name="B",
                success=True,
                records=[],
                usage=TokenUsage(llm_calls=2, input_tokens=200, output_tokens=80, total_tokens=280),
            ),
        })
        usage = result.total_usage
        assert usage.llm_calls == 3
        assert usage.input_tokens == 300
        assert usage.total_tokens == 430

    def test_total_usage_skips_none(self):
        result = WorkbookResult(results={
            "A": SheetResult(sheet_name="A", success=False, error="err"),
            "B": SheetResult(
                sheet_name="B",
                success=True,
                records=[],
                usage=TokenUsage(llm_calls=1, input_tokens=50, output_tokens=25, total_tokens=75),
            ),
        })
        assert result.total_usage.llm_calls == 1

    def test_getitem(self):
        sheet = SheetResult(
            sheet_name="Sales",
            success=True,
            records=[SalesRow(product="Widget", revenue=1000.0)],
        )
        result = WorkbookResult(results={"Sales": sheet})
        assert result["Sales"].sheet_name == "Sales"
        assert result["Sales"].records[0].product == "Widget"

    def test_contains(self):
        result = WorkbookResult(results={
            "Sales": SheetResult(sheet_name="Sales", success=True, records=[]),
        })
        assert "Sales" in result
        assert "Missing" not in result

    def test_iteration(self):
        result = WorkbookResult(results={
            "Sales": SheetResult(sheet_name="Sales", success=True, records=[]),
            "Expenses": SheetResult(sheet_name="Expenses", success=True, records=[]),
        })
        names = list(result)
        assert names == ["Sales", "Expenses"]

    def test_getitem_keyerror(self):
        result = WorkbookResult(results={})
        with pytest.raises(KeyError):
            _ = result["Nonexistent"]


# * Extractor integration tests

class TestExtractWorkbook:
    async def test_all_sheets_succeed(self, multi_sheet_xlsx):
        extractor = Extractor()

        async def mock_extract(encoded: Any, schema: Any, instructions: Any, **kw: Any) -> list:
            if schema is SalesRow:
                return [SalesRow(product="Widget", revenue=1000.0)]
            if schema is ExpenseRow:
                return [ExpenseRow(category="Office", amount=500.0)]
            return [InventoryRow(item="Widget", quantity=100)]

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            side_effect=mock_extract,
        ):
            result = await extractor.extract_workbook(
                multi_sheet_xlsx,
                sheet_schemas={
                    "Sales": SalesRow,
                    "Expenses": ExpenseRow,
                    "Inventory": InventoryRow,
                },
            )

        assert result.succeeded == 3
        assert result.failed == 0
        assert result["Sales"].records[0].product == "Widget"
        assert result["Expenses"].records[0].category == "Office"
        assert result["Inventory"].records[0].quantity == 100

    async def test_partial_failure(self, multi_sheet_xlsx):
        extractor = Extractor()

        call_count = 0

        async def mock_extract(encoded: Any, schema: Any, instructions: Any, **kw: Any) -> list:
            nonlocal call_count
            call_count += 1
            if schema is ExpenseRow:
                raise ValueError("LLM error")
            if schema is SalesRow:
                return [SalesRow(product="Widget", revenue=1000.0)]
            return [InventoryRow(item="Widget", quantity=100)]

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            side_effect=mock_extract,
        ):
            result = await extractor.extract_workbook(
                multi_sheet_xlsx,
                sheet_schemas={
                    "Sales": SalesRow,
                    "Expenses": ExpenseRow,
                    "Inventory": InventoryRow,
                },
            )

        assert result.succeeded == 2
        assert result.failed == 1
        assert not result["Expenses"].success
        assert "ValueError" in result["Expenses"].error

    async def test_missing_sheet(self, multi_sheet_xlsx):
        extractor = Extractor()

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            new_callable=AsyncMock,
            return_value=[SalesRow(product="X", revenue=1.0)],
        ):
            result = await extractor.extract_workbook(
                multi_sheet_xlsx,
                sheet_schemas={
                    "Sales": SalesRow,
                    "NonExistent": ExpenseRow,
                },
            )

        assert result.succeeded == 1
        assert result.failed == 1
        assert not result["NonExistent"].success
        assert "not found" in result["NonExistent"].error

    async def test_empty_schema_dict(self, multi_sheet_xlsx):
        extractor = Extractor()
        result = await extractor.extract_workbook(
            multi_sheet_xlsx,
            sheet_schemas={},
        )
        assert result.total == 0
        assert result.succeeded == 0

    async def test_results_keyed_by_sheet_name(self, multi_sheet_xlsx):
        extractor = Extractor()

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            new_callable=AsyncMock,
            return_value=[],
        ):
            result = await extractor.extract_workbook(
                multi_sheet_xlsx,
                sheet_schemas={
                    "Sales": SalesRow,
                    "Inventory": InventoryRow,
                },
            )

        assert set(result.sheet_names) == {"Sales", "Inventory"}


class TestExtractWorkbookSync:
    def test_sync_wrapper(self, multi_sheet_xlsx):
        extractor = Extractor()

        with patch(
            "xlstruct.extraction.engine.ExtractionEngine.extract",
            new_callable=AsyncMock,
            return_value=[SalesRow(product="X", revenue=1.0)],
        ):
            result = extractor.extract_workbook_sync(
                multi_sheet_xlsx,
                sheet_schemas={"Sales": SalesRow},
            )

        assert result.succeeded == 1
        assert result["Sales"].success
