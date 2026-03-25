"""Tests for strict_formulas configuration toggle."""

import io
import logging

import openpyxl
import pytest

from xlstruct.config import ExtractorConfig
from xlstruct.encoder.compressed import CompressedEncoder
from xlstruct.exceptions import ReaderError
from xlstruct.reader.hybrid_reader import HybridReader
from xlstruct.schemas.core import CellData, SheetData

# * Fixtures


@pytest.fixture
def xlsx_with_uncached_formulas() -> bytes:
    """Generate a .xlsx file where formula cells have no cached value.

    Simulates Google Sheets exports: openpyxl writes formula strings
    but calamine sees no computed value for those cells.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # ^ Header
    ws["A1"] = "Item"
    ws["B1"] = "Qty"
    ws["C1"] = "Price"
    ws["D1"] = "Total"

    # ^ Data row with values
    ws["A2"] = "Widget"
    ws["B2"] = 10
    ws["C2"] = 25.0
    # ^ Formula cell — openpyxl stores the formula string but no cached value
    ws["D2"] = "=B2*C2"

    ws["A3"] = "Gadget"
    ws["B3"] = 5
    ws["C3"] = 42.0
    ws["D3"] = "=B3*C3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def sheet_with_uncached_formulas() -> SheetData:
    """A SheetData with formula cells that have no cached value."""
    cells = [
        CellData(row=1, col=1, value="Item", data_type="s"),
        CellData(row=1, col=2, value="Qty", data_type="s"),
        CellData(row=1, col=3, value="Price", data_type="s"),
        CellData(row=1, col=4, value="Total", data_type="s"),
        CellData(row=2, col=1, value="Widget", data_type="s"),
        CellData(row=2, col=2, value=10, data_type="n"),
        CellData(row=2, col=3, value=25.0, data_type="n"),
        # ^ Formula cell with no cached_value — simulates Google Sheets export
        CellData(
            row=2,
            col=4,
            value="=B2*C2",
            formula="=B2*C2",
            cached_value=None,
            data_type="f",
        ),
    ]
    return SheetData(
        name="Sales",
        dimensions="A1:D2",
        cells=cells,
        merged_ranges=[],
        row_count=2,
        col_count=4,
    )


# * _check_uncached_formulas unit tests


class TestCheckUncachedFormulas:
    """Test _check_uncached_formulas with strict toggle."""

    def test_strict_true_raises_reader_error(self, sheet_with_uncached_formulas: SheetData):
        """Default behavior: strict=True raises ReaderError."""
        with pytest.raises(ReaderError, match="formula cell"):
            HybridReader._check_uncached_formulas(sheet_with_uncached_formulas, strict=True)

    def test_strict_false_logs_warning(
        self, sheet_with_uncached_formulas: SheetData, caplog: pytest.LogCaptureFixture
    ):
        """strict=False should log warning instead of raising."""
        with caplog.at_level(logging.WARNING, logger="xlstruct.reader.hybrid_reader"):
            HybridReader._check_uncached_formulas(sheet_with_uncached_formulas, strict=False)

        assert len(caplog.records) == 1
        assert "formula cell" in caplog.records[0].message
        assert "D2" in caplog.records[0].message

    def test_no_uncached_formulas_no_error(self):
        """No uncached formulas — should pass silently in both modes."""
        cells = [
            CellData(row=1, col=1, value="A", data_type="s"),
            CellData(
                row=1,
                col=2,
                value="=A1",
                formula="=A1",
                cached_value="A",
                data_type="f",
            ),
        ]
        sheet = SheetData(name="OK", cells=cells, row_count=1, col_count=2)
        # ^ Neither strict mode should raise or warn
        HybridReader._check_uncached_formulas(sheet, strict=True)
        HybridReader._check_uncached_formulas(sheet, strict=False)


# * HybridReader.read() integration tests


class TestStrictFormulasReader:
    """Test strict_formulas parameter on HybridReader.read()."""

    def test_strict_true_raises_for_uncached(self, xlsx_with_uncached_formulas: bytes):
        """strict_formulas=True (default) raises ReaderError for uncached formulas."""
        reader = HybridReader()
        with pytest.raises(ReaderError, match="formula cell"):
            reader.read(xlsx_with_uncached_formulas, strict_formulas=True)

    def test_strict_false_returns_workbook(self, xlsx_with_uncached_formulas: bytes):
        """strict_formulas=False returns WorkbookData without raising."""
        reader = HybridReader()
        wb = reader.read(xlsx_with_uncached_formulas, strict_formulas=False)

        assert len(wb.sheets) == 1
        assert wb.sheets[0].name == "Sales"

    def test_strict_false_formula_strings_in_display_value(
        self, xlsx_with_uncached_formulas: bytes
    ):
        """When strict_formulas=False, formula cells fall back to formula string."""
        reader = HybridReader()
        wb = reader.read(xlsx_with_uncached_formulas, strict_formulas=False)
        sheet = wb.sheets[0]

        # ^ D2 and D3 are formula cells with no cached value
        d2 = sheet.get_cell(2, 4)
        assert d2 is not None
        assert d2.formula is not None
        assert d2.formula.startswith("=")
        # ^ display_value falls back to value (the formula string)
        assert d2.display_value == d2.value
        assert isinstance(d2.display_value, str)
        assert d2.display_value.startswith("=")

    def test_default_is_strict(self, xlsx_with_uncached_formulas: bytes):
        """Default behavior (no explicit strict_formulas) should be strict."""
        reader = HybridReader()
        with pytest.raises(ReaderError):
            reader.read(xlsx_with_uncached_formulas)


# * ExtractorConfig tests


class TestStrictFormulasConfig:
    """Test strict_formulas field on ExtractorConfig."""

    def test_default_is_true(self):
        config = ExtractorConfig()
        assert config.strict_formulas is True

    def test_can_set_false(self):
        config = ExtractorConfig(strict_formulas=False)
        assert config.strict_formulas is False


# * Encoder integration test


class TestStrictFormulasEncoder:
    """Test that formula strings appear in encoded output when strict_formulas=False."""

    def test_formula_strings_in_encoded_output(self, xlsx_with_uncached_formulas: bytes):
        """Encoded output should contain formula strings when strict_formulas=False."""
        reader = HybridReader()
        wb = reader.read(xlsx_with_uncached_formulas, strict_formulas=False)
        sheet = wb.sheets[0]

        encoder = CompressedEncoder()
        encoded = encoder.encode(sheet)

        # ^ The encoded output should contain at least one formula string
        assert "=B2*C2" in encoded or "=B3*C3" in encoded
