"""Tests for optional formula evaluation via formulas library."""

import io
import logging
import sys
from unittest.mock import patch

import openpyxl
import pytest

from xlstruct.reader.hybrid_reader import HybridReader
from xlstruct.schemas.core import CellData, SheetData

# * Fixtures


@pytest.fixture
def sheet_with_sum() -> SheetData:
    """SheetData with a SUM formula and supporting value cells."""
    return SheetData(
        name="Sheet1",
        dimensions="A1:B3",
        row_count=3,
        col_count=2,
        cells=[
            CellData(row=1, col=1, value=10, cached_value=10, data_type="n"),
            CellData(row=2, col=1, value=20, cached_value=20, data_type="n"),
            CellData(
                row=3,
                col=1,
                value="=SUM(A1:A2)",
                formula="=SUM(A1:A2)",
                cached_value=None,
                data_type="f",
            ),
        ],
    )


@pytest.fixture
def sheet_no_formulas() -> SheetData:
    """SheetData with no formula cells."""
    return SheetData(
        name="Sheet1",
        dimensions="A1:A2",
        row_count=2,
        col_count=1,
        cells=[
            CellData(row=1, col=1, value="Hello", cached_value="Hello", data_type="s"),
            CellData(row=2, col=1, value=42, cached_value=42, data_type="n"),
        ],
    )


@pytest.fixture
def xlsx_with_formula() -> bytes:
    """Build a real .xlsx file with a SUM formula (no cached value)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# * Import error test


def test_import_error_when_formulas_not_installed():
    """Verify helpful ImportError when formulas package is missing."""
    from xlstruct.reader.formula_eval import _import_formulas

    with patch.dict(sys.modules, {"formulas": None}):
        with pytest.raises(ImportError, match="pip install xlstruct\\[formulas\\]"):
            _import_formulas()


# * Default skip test


def test_no_formulas_returns_unchanged(sheet_no_formulas):
    """When no formula cells exist, evaluate_sheet_formulas returns the sheet unchanged."""
    from xlstruct.reader.formula_eval import evaluate_sheet_formulas

    result = evaluate_sheet_formulas(sheet_no_formulas)
    assert result.cells == sheet_no_formulas.cells


# * Basic SUM evaluation


def test_evaluate_sum_formula(sheet_with_sum):
    """SUM(A1:A2) with values 10 and 20 should evaluate to 30."""
    from xlstruct.reader.formula_eval import evaluate_sheet_formulas

    result = evaluate_sheet_formulas(sheet_with_sum)

    # ^ Find the formula cell (row=3, col=1)
    formula_cell = next(c for c in result.cells if c.row == 3 and c.col == 1)
    assert formula_cell.cached_value == pytest.approx(30)
    assert formula_cell.formula == "=SUM(A1:A2)"


# * Failed evaluation warning


def test_failed_eval_logs_warning(sheet_with_sum, caplog):
    """A cell with an unparseable formula result key should log a warning."""
    from xlstruct.reader.formula_eval import evaluate_sheet_formulas

    # ^ Patch formulas to return a result with a bad key
    class FakeModel:
        def loads(self, path):
            return self

        def finish(self):
            return self

        def calculate(self):
            return {"INVALID_KEY_NO_BANG": 999}

    class FakeFormulas:
        @staticmethod
        def ExcelModel():
            return FakeModel()

    with patch.dict(sys.modules, {"formulas": FakeFormulas}):
        with caplog.at_level(logging.WARNING, logger="xlstruct.reader.formula_eval"):
            result = evaluate_sheet_formulas(sheet_with_sum)

    # ^ Formula cell should remain unchanged (no computed value matched)
    formula_cell = next(c for c in result.cells if c.row == 3 and c.col == 1)
    assert formula_cell.cached_value is None
    assert "Failed to parse formula result key" in caplog.text


# * Integration: HybridReader with evaluate_formulas=True


def test_hybrid_reader_evaluate_formulas_integration(xlsx_with_formula):
    """HybridReader.read() with evaluate_formulas=True should compute formula values."""
    reader = HybridReader()
    workbook = reader.read(xlsx_with_formula, source_ext=".xlsx", evaluate_formulas=True)

    sheet = workbook.sheets[0]
    formula_cell = next(
        (c for c in sheet.cells if c.formula),
        None,
    )
    assert formula_cell is not None
    assert formula_cell.cached_value == pytest.approx(30)


# * Integration: evaluate_formulas=False (default) preserves existing behavior


def test_hybrid_reader_default_no_eval(xlsx_with_formula):
    """HybridReader.read() without evaluate_formulas should not evaluate formulas.

    The .xlsx fixture was created with openpyxl which does not write cached values
    for formulas, so the formula cell will have a cached_value from calamine
    (which reads the calculated value if Excel saved it). Since openpyxl-generated
    files have no cached value, this will raise ReaderError for uncached formulas.
    """
    from xlstruct.exceptions import ReaderError

    reader = HybridReader()
    with pytest.raises(ReaderError, match="no cached value"):
        reader.read(xlsx_with_formula, source_ext=".xlsx", evaluate_formulas=False)
