"""Tests for CLI module."""

import json
import sys
from unittest.mock import MagicMock, patch

import pytest
from click.exceptions import Exit as ClickExit
from typer.testing import CliRunner

from xlstruct.cli import _resolve_schema, app

runner = CliRunner()


class TestResolveSchema:
    def test_valid_schema(self):
        cls = _resolve_schema("pydantic:BaseModel")
        from pydantic import BaseModel

        assert cls is BaseModel

    def test_missing_colon(self):
        with pytest.raises(ClickExit):
            _resolve_schema("no_colon_here")

    def test_invalid_module(self):
        with pytest.raises(ClickExit):
            _resolve_schema("nonexistent_module_xyz:Foo")

    def test_invalid_class(self):
        with pytest.raises(ClickExit):
            _resolve_schema("pydantic:NonExistentClassName")


class TestExtractCommand:
    def test_missing_schema_option(self):
        """CLI should fail without --schema."""
        result = runner.invoke(app, ["test.xlsx"])
        assert result.exit_code != 0

    def test_extract_with_mocked_extractor(self, tmp_path):
        """Full CLI flow with mocked Extractor."""
        from pydantic import BaseModel

        class Item(BaseModel):
            name: str
            value: int

        mock_results = [Item(name="Test", value=42)]

        # ^ Write a dummy schema module
        schema_file = tmp_path / "schema.py"
        schema_file.write_text(
            "from pydantic import BaseModel\n\n"
            "class Item(BaseModel):\n"
            "    name: str\n"
            "    value: int\n"
        )

        # ^ Create a dummy xlsx
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "Test"
        ws["B2"] = 42
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_path = tmp_path / "test.xlsx"
        xlsx_path.write_bytes(buf.getvalue())

        with (
            patch("sys.path", [str(tmp_path)] + sys.path),
            patch("xlstruct.extractor.Extractor.extract_sync", return_value=mock_results),
            patch(
                "xlstruct.extraction.engine.ExtractionEngine._build_client",
                return_value=MagicMock(),
            ),
        ):
            result = runner.invoke(app, [
                str(xlsx_path),
                "--schema", "schema:Item",
            ])

        assert result.exit_code == 0, result.output
        output = json.loads(result.output)
        assert len(output) == 1
        assert output[0]["name"] == "Test"
        assert output[0]["value"] == 42

    def test_extract_output_to_file(self, tmp_path):
        """Test --output flag writes JSON to file."""
        from pydantic import BaseModel

        class Row(BaseModel):
            x: int

        mock_results = [Row(x=1), Row(x=2)]

        # ^ Schema module
        schema_file = tmp_path / "s.py"
        schema_file.write_text(
            "from pydantic import BaseModel\n\n"
            "class Row(BaseModel):\n"
            "    x: int\n"
        )

        # ^ Dummy xlsx
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "X"
        ws["A2"] = 1
        ws["A3"] = 2
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_path = tmp_path / "data.xlsx"
        xlsx_path.write_bytes(buf.getvalue())
        output_path = tmp_path / "out.json"

        with (
            patch("sys.path", [str(tmp_path)] + sys.path),
            patch("xlstruct.extractor.Extractor.extract_sync", return_value=mock_results),
            patch(
                "xlstruct.extraction.engine.ExtractionEngine._build_client",
                return_value=MagicMock(),
            ),
        ):
            result = runner.invoke(app, [
                str(xlsx_path),
                "--schema", "s:Row",
                "--output", str(output_path),
            ])

        assert result.exit_code == 0, result.output
        assert output_path.exists()
        data = json.loads(output_path.read_text())
        assert len(data) == 2
