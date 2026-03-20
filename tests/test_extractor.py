"""Integration tests for Extractor class with mocked LLM extraction."""

import io
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest
from pydantic import BaseModel

from xlstruct.config import ExtractionMode, ExtractorConfig
from xlstruct.extractor import ExtractionResult, Extractor
from xlstruct.schemas.report import ExtractionReport
from xlstruct.schemas.usage import TokenUsage

# * Test schemas

class Product(BaseModel):
    name: str
    price: float
    stock: int


# * Fixtures

@pytest.fixture(autouse=True)
def _mock_instructor():
    """Mock build_instructor_client to avoid real API calls."""
    with patch(
        "xlstruct.extraction.engine.build_instructor_client",
        return_value=MagicMock(),
    ):
        yield


@pytest.fixture
def product_xlsx_bytes() -> bytes:
    """Real xlsx with product data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws["A1"] = "Name"
    ws["B1"] = "Price"
    ws["C1"] = "Stock"
    ws["A2"] = "Apple"
    ws["B2"] = 1.5
    ws["C2"] = 100
    ws["A3"] = "Banana"
    ws["B3"] = 0.75
    ws["C3"] = 200
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def product_xlsx_file(product_xlsx_bytes: bytes, tmp_path) -> str:
    """Write product xlsx to a temp file and return path."""
    path = tmp_path / "products.xlsx"
    path.write_bytes(product_xlsx_bytes)
    return str(path)


@pytest.fixture
def multi_sheet_xlsx_file(tmp_path) -> str:
    """Xlsx with two sheets."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Name"
    ws1["B1"] = "Value"
    ws1["A2"] = "X"
    ws1["B2"] = 10

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Name"
    ws2["B1"] = "Value"
    ws2["A2"] = "Y"
    ws2["B2"] = 20

    buf = io.BytesIO()
    wb.save(buf)
    path = tmp_path / "multi.xlsx"
    path.write_bytes(buf.getvalue())
    return str(path)


# * Tests

class TestExtractorInit:
    def test_default_init(self):
        extractor = Extractor()
        assert extractor._config.provider == "anthropic/claude-sonnet-4-6"

    def test_custom_provider(self):
        extractor = Extractor(provider="anthropic/claude-sonnet-4-20250514")
        assert extractor._config.provider == "anthropic/claude-sonnet-4-20250514"

    def test_with_config(self):
        config = ExtractorConfig(provider="openai/gpt-4o", temperature=0.5)
        extractor = Extractor(config=config)
        assert extractor._config.temperature == 0.5

    def test_kwargs_forwarded(self):
        extractor = Extractor(provider="openai/gpt-4o", temperature=0.7, max_retries=5)
        assert extractor._config.temperature == 0.7
        assert extractor._config.max_retries == 5


class TestExtractorExtract:
    @pytest.mark.asyncio
    async def test_extract_single_sheet(self, product_xlsx_file):
        """Test full pipeline with mocked LLM engine."""
        expected = [
            Product(name="Apple", price=1.5, stock=100),
            Product(name="Banana", price=0.75, stock=200),
        ]

        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = expected
            results = await extractor.extract(product_xlsx_file, Product)

        assert len(results) == 2
        assert results[0].name == "Apple"
        assert results[1].price == 0.75
        mock_extract.assert_called_once()

    @pytest.mark.asyncio
    async def test_extract_with_instructions(self, product_xlsx_file):
        """Test that instructions are passed through to the engine."""
        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = []
            await extractor.extract(
                product_xlsx_file, Product, instructions="Extract all products"
            )

        # ^ Verify instructions argument was passed
        call_args = mock_extract.call_args
        assert call_args[0][2] == "Extract all products"

    @pytest.mark.asyncio
    async def test_extract_with_sheet_name(self, multi_sheet_xlsx_file):
        """Test targeting a specific sheet."""
        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = []
            await extractor.extract(multi_sheet_xlsx_file, Product, sheet="Sheet2")

        mock_extract.assert_called_once()

class TestExtractorExtractSync:
    def test_extract_sync(self, product_xlsx_file):
        """Test synchronous wrapper."""
        expected = [Product(name="Apple", price=1.5, stock=100)]

        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = expected
            results = extractor.extract_sync(product_xlsx_file, Product)

        assert len(results) == 1
        assert results[0].name == "Apple"


class TestExtractorPipeline:
    @pytest.mark.asyncio
    async def test_encoder_selected_based_on_size(self, product_xlsx_file):
        """Small sheet → CompressedEncoder should be used."""
        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = []
            await extractor.extract(product_xlsx_file, Product)

        # ^ Verify encoded text was markdown format (starts with ## Sheet:)
        encoded_text = mock_extract.call_args[0][0]
        assert '## Sheet: "Products"' in encoded_text

    @pytest.mark.asyncio
    async def test_chunking_not_triggered_for_small_sheet(self, product_xlsx_file):
        """Small sheet should not trigger chunking."""
        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = []
            await extractor.extract(product_xlsx_file, Product)

        # ^ Only one call = no chunking
        mock_extract.assert_called_once()



class TestExtractionResult:
    @pytest.fixture
    def _empty_report(self) -> ExtractionReport:
        return ExtractionReport(mode=ExtractionMode.DIRECT, usage=TokenUsage())

    def test_report_mode_default(self, _empty_report):
        """report.mode reflects the extraction mode."""
        result = ExtractionResult([], report=_empty_report)
        assert result.report.mode == ExtractionMode.DIRECT

    def test_report_mode_codegen(self):
        """report.mode can be set to CODEGEN."""
        report = ExtractionReport(mode=ExtractionMode.CODEGEN, usage=TokenUsage())
        result = ExtractionResult([], report=report)
        assert result.report.mode == ExtractionMode.CODEGEN

    def test_report_source_rows_default_empty(self, _empty_report):
        """report.source_rows defaults to empty list."""
        result = ExtractionResult([], report=_empty_report)
        assert result.report.source_rows == []

    def test_report_source_rows_set(self):
        """report.source_rows can be set explicitly."""
        rows = [[2], [3], [4, 5]]
        report = ExtractionReport(
            mode=ExtractionMode.DIRECT, usage=TokenUsage(), source_rows=rows
        )
        result = ExtractionResult([], report=report)
        assert result.report.source_rows == [[2], [3], [4, 5]]

    def test_report_summary(self):
        """report.summary() returns a human-readable string."""
        report = ExtractionReport(
            mode=ExtractionMode.DIRECT,
            usage=TokenUsage(input_tokens=100, output_tokens=50, total_tokens=150),
            source_rows=[[2], [3]],
        )
        text = report.summary()
        assert "direct" in text
        assert "150" in text
        assert "2 records mapped" in text

    def test_to_dataframe(self, _empty_report):
        """to_dataframe() converts records to a pandas DataFrame."""
        pd = pytest.importorskip("pandas")
        items = [
            Product(name="Apple", price=1.5, stock=100),
            Product(name="Banana", price=0.75, stock=200),
        ]
        result = ExtractionResult(items, report=_empty_report)
        df = result.to_dataframe()

        assert isinstance(df, pd.DataFrame)
        assert len(df) == 2
        assert list(df.columns) == ["name", "price", "stock"]
        assert df.iloc[0]["name"] == "Apple"
        assert df.iloc[1]["price"] == 0.75

    def test_to_dataframe_empty(self, _empty_report):
        """to_dataframe() returns empty DataFrame for empty results."""
        pd = pytest.importorskip("pandas")
        result: ExtractionResult[Product] = ExtractionResult([], report=_empty_report)
        df = result.to_dataframe()

        assert isinstance(df, pd.DataFrame)
        assert len(df) == 0

    def test_to_dataframe_import_error(self, _empty_report):
        """to_dataframe() raises ImportError when pandas is missing."""
        result = ExtractionResult(
            [Product(name="A", price=1.0, stock=1)], report=_empty_report
        )
        with patch.dict("sys.modules", {"pandas": None}):
            with pytest.raises(ImportError, match="pandas is required"):
                result.to_dataframe()


class TestExtractionResultFromExtractor:
    @pytest.mark.asyncio
    async def test_legacy_extract_report(self, product_xlsx_file):
        """Legacy schema-based extraction should produce report with mode=DIRECT."""
        extractor = Extractor()
        with patch.object(extractor._engine, "extract", new_callable=AsyncMock) as mock_extract:
            mock_extract.return_value = [Product(name="Apple", price=1.5, stock=100)]
            result = await extractor.extract(product_xlsx_file, Product)

        assert result.report.mode == ExtractionMode.DIRECT
        assert result.report.usage.llm_calls == 0  # ^ mocked, no real calls
        assert result.report.source_rows == []


class TestExtractorLoadWorkbook:
    @pytest.mark.asyncio
    async def test_file_metadata(self, product_xlsx_file):
        """Verify file_name and file_size are set."""
        extractor = Extractor()
        workbook = await extractor._load_workbook(product_xlsx_file)

        assert workbook.file_name == "products.xlsx"
        assert workbook.file_size > 0

    @pytest.mark.asyncio
    async def test_storage_options_merged(self, product_xlsx_file):
        """Verify config storage_options and per-call options are merged."""
        config = ExtractorConfig(storage_options={"anon": True})
        extractor = Extractor(config=config)

        workbook = await extractor._load_workbook(product_xlsx_file)
        assert len(workbook.sheets) == 1
